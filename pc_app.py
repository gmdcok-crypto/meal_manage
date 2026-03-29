import sys
import warnings
# Suppress PyQt5 sip deprecation warnings
warnings.filterwarnings("ignore", category=DeprecationWarning, message=".*sipPyTypeDict.*")
import json
import httpx
import asyncio
import websockets
import threading
from datetime import datetime, date, timedelta, timezone
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QStackedWidget, QTableWidget, QTableWidgetItem,
    QLineEdit, QComboBox, QFrame, QHeaderView, QGraphicsDropShadowEffect,
    QAbstractItemView, QDialog, QFormLayout, QMessageBox, QInputDialog, QStyledItemDelegate, QTimeEdit, QDateEdit, QCalendarWidget, QLayout,     QPlainTextEdit, QSizePolicy, QCheckBox, QSpinBox, QGroupBox, QScrollArea, QDialogButtonBox,
)
from PyQt5.QtCore import Qt, QSize, pyqtSignal, QTimer, QThread, QTime, QDate, QByteArray
from PyQt5.QtGui import QFont, QColor, QIcon, QPixmap, QResizeEvent, QIntValidator

# --- Config (override via env or edit for deployment) ---
# 기본: Railway 백엔드 (가이드 예시 주소). 로컬 사용 시 MEAL_API_BASE_URL 으로 http://localhost:8000/api/admin 설정.
import os
_DEFAULT_RAILWAY = "https://web-production-e758d.up.railway.app/api/admin"
_API_BASE = os.environ.get("MEAL_API_BASE_URL", _DEFAULT_RAILWAY)
API_BASE_URL = _API_BASE
# PWA 공지사항 저장 경로 (프로젝트 static 폴더)
_MEAL_MANAGE_ROOT = os.path.dirname(os.path.abspath(__file__))
NOTICE_HTML_PATH = os.path.join(_MEAL_MANAGE_ROOT, "static", "notice.html")
_ws_origin = _API_BASE.replace("https://", "wss://").replace("http://", "ws://").split("/api")[0]
WS_URL = _ws_origin + "/api/admin/ws"
API_TIMEOUT = 10.0
# 로그인: 기본은 창 생략 후 메인 바로 표시. 로그인 창을 쓰려면 환경변수 MEAL_PC_REQUIRE_LOGIN=1 (또는 true/yes)
PC_APP_SKIP_LOGIN = os.environ.get("MEAL_PC_REQUIRE_LOGIN", "").strip().lower() not in ("1", "true", "yes")


def _run_print_and_qlight(meal_data: dict, device_settings: dict):
    """백그라운드 스레드에서 프린터·경광등 신호 전송 (PC 앱과 같은 망의 장치용)."""
    if not device_settings:
        return
    # 프로젝트 루트에서 실행하지 않아도 모듈 로드되도록
    if _MEAL_MANAGE_ROOT not in sys.path:
        sys.path.insert(0, _MEAL_MANAGE_ROOT)
    emp_no = (meal_data.get("emp_no") or "").strip()
    name = (meal_data.get("name") or "").strip()
    meal_type_label = (meal_data.get("meal_type_label") or "").strip()
    date_time_str = (meal_data.get("date_time_str") or "").strip()
    # 프린터
    if device_settings.get("printer_enabled") and (device_settings.get("printer_host") or "").strip():
        try:
            import bixolon_print
            bixolon_print.print_image_only(
                host=(device_settings.get("printer_host") or "").strip(),
                port=int(device_settings.get("printer_port") or 9100),
                stored_image_number=int(device_settings.get("printer_stored_image_number") or 1),
                emp_no=emp_no,
                name=name,
                date_time_str=date_time_str,
                meal_type=meal_type_label,
            )
        except Exception:
            pass
    # 경광등
    if device_settings.get("qlight_enabled") and (device_settings.get("qlight_host") or "").strip():
        try:
            import qlight_st45l
            q_host = (device_settings.get("qlight_host") or "").strip()
            q_port = int(device_settings.get("qlight_port") or 20000)
            ok = qlight_st45l.trigger_ok(
                host=q_host,
                port=q_port,
                blink=False,
            )
            print(f"[QLIGHT] trigger_ok host={q_host} port={q_port} result={ok}")
        except Exception as e:
            print(f"[QLIGHT] trigger exception: {e}")


# --- WebSocket Client Thread ---
class WSClient(QThread):
    message_received = pyqtSignal(dict)
    
    def __init__(self, ws_url=None):
        super().__init__()
        self.ws_url = ws_url or WS_URL
        self.running = True
        self.loop = None
        
    def run(self):
        self.loop = asyncio.new_event_loop()
        asyncio.set_event_loop(self.loop)
        self.loop.run_until_complete(self.listen())
        
    async def listen(self):
        while self.running:
            try:
                print(f"[WS] connecting: {self.ws_url}")
                async with websockets.connect(self.ws_url) as ws:
                    self.ws = ws
                    print("[WS] connected")
                    while self.running:
                        msg = await ws.recv()
                        try:
                            data = json.loads(msg)
                            self.message_received.emit(data)
                        except (json.JSONDecodeError, TypeError):
                            pass
            except Exception as e:
                # Reconnect on error after delay
                print(f"[WS] error/reconnect: {e}")
                if self.running:
                    await asyncio.sleep(3)
                    
    def stop(self):
        self.running = False
        if hasattr(self, 'ws') and self.ws and self.loop is not None:
            try:
                asyncio.run_coroutine_threadsafe(self.ws.close(), self.loop)
            except Exception:
                pass
        self.quit()
        self.wait(3000)

# --- Worker Thread for Async Data Loading ---
class DataLoader(QThread):
    finished = pyqtSignal(object)
    error = pyqtSignal(str)

    def __init__(self, func, *args, **kwargs):
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs

    def run(self):
        try:
            result = self.func(*self.args, **self.kwargs)
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


def _normalize_api_list_payload(body):
    """GET 등 응답이 최상위 배열이 아닐 때(래핑 JSON) 리스트만 추출."""
    if isinstance(body, list):
        return body
    if isinstance(body, dict):
        for key in ("items", "results", "data", "policies", "rows"):
            v = body.get(key)
            if isinstance(v, list):
                return v
    return []


# --- Premium QSS Styling ---
QSS = """
QMainWindow { background-color: #0f172a; font-family: 'Malgun Gothic', 'Segoe UI', sans-serif; font-weight: bold; }
QWidget#Sidebar { background-color: #1e293b; border-right: 1px solid #334155; min-width: 260px; }
QLabel#Logo { color: #f8fafc; font-size: 28px; font-weight: bold; font-family: 'Malgun Gothic'; padding: 25px; }
QPushButton#MenuBtn {
    background-color: transparent; color: #94a3b8; border: none; text-align: left;
    padding: 15px 25px; font-size: 20px; font-weight: bold; font-family: 'Malgun Gothic'; border-radius: 8px; margin: 4px 12px;
}
QPushButton#MenuBtn:hover { background-color: #334155; color: #f8fafc; }
QPushButton#MenuBtn[active="true"] { background-color: #6366f1; color: #ffffff; }
QWidget#ContentArea { background-color: #0f172a; }
QStackedWidget { background: transparent; }
/* 각 탭 루트(설정·원시데이터 등) 바탕 = ContentArea와 동일. 앱 전역 QSS라 자식 테이블/버튼 QSS와 충돌 없음 */
QStackedWidget > QWidget { background-color: #0f172a; }
QLabel#HeaderTitle { color: #f8fafc; font-size: 36px; font-weight: bold; font-family: 'Malgun Gothic'; background-color: transparent; }
QWidget#NoticeScreen { background-color: #0f172a; }
QWidget#NoticeScreen QLabel#HeaderTitle { margin: 0; padding: 0; }
QWidget#NoticeScreen QLabel#NoticeHint { margin: 0; padding: 0; }
/* 설정 탭 바탕 = 공지사항 탭(NoticeScreen)과 동일 색. 앱 전역 QSS만 사용(자식 테이블·버튼 QSS 유지) */
QWidget#SettingsScreenRoot { background-color: #0f172a; }
/* 설정 탭 QScrollArea 뷰포트·내부는 기본 흰색 → 테이블 아래/위 빈 영역이 하얗게 보임 */
QScrollArea#SettingsScroll { background-color: #0f172a; border: none; }
QWidget#SettingsScrollViewport { background-color: #0f172a; }
QWidget#SettingsScrollInner { background-color: #0f172a; }
QWidget#SettingsScrollInner QLabel { background-color: transparent; }
QFrame#StatCard { background-color: #1e293b; border-radius: 12px; border: 1px solid #334155; padding: 15px; }
QLabel#StatValue { color: #f8fafc; font-size: 52px; font-weight: bold; font-family: 'Malgun Gothic'; }
QLabel#StatLabel { color: #94a3b8; font-size: 22px; font-weight: bold; font-family: 'Malgun Gothic'; }
QTableWidget { background-color: #1e293b; color: #f1f5f9; gridline-color: #334155; border: 1px solid #334155; border-radius: 12px; alternate-background-color: #1a2333; font-size: 17px; selection-background-color: #3b82f6; outline: none; font-weight: bold; }
QTableWidget::item { padding: 0px; border-bottom: 1px solid #334155; border-right: 1px solid #334155; }
QTableWidget::item:selected { background-color: #3b82f6; color: #ffffff; font-weight: bold; border-bottom: 1px solid #334155; border-right: 1px solid #334155; }
QHeaderView::section { background-color: #111b2d; color: #94a3b8; padding: 12px; font-weight: bold; font-size: 19px; font-family: 'Malgun Gothic'; border: none; border-right: 1px solid #334155; border-bottom: 2px solid #334155; }
QLineEdit, QTimeEdit { background-color: #1e293b; border: 1px solid #475569; border-radius: 8px; color: #f8fafc; padding: 10px 15px; font-size: 19px; height: 40px; font-weight: bold; font-family: 'Malgun Gothic'; }
QDateEdit { background-color: #1e293b; border: 1px solid #475569; border-radius: 8px; color: #f8fafc; padding-left: 10px; padding-right: 30px; font-size: 19px; height: 40px; font-weight: bold; font-family: 'Malgun Gothic'; }
QDateEdit::up-button { subcontrol-origin: border; subcontrol-position: top right; width: 20px; height: 16px; border-left: 1px solid #475569; border-top-right-radius: 8px; background-color: #334155; }
QDateEdit::down-button { subcontrol-origin: border; subcontrol-position: bottom right; width: 20px; height: 16px; border-left: 1px solid #475569; border-top: 1px solid #475569; border-bottom-right-radius: 8px; background-color: #334155; }
QDateEdit::drop-down { subcontrol-origin: border; subcontrol-position: top right; width: 26px; border-left: 1px solid #475569; border-top-right-radius: 8px; border-bottom-right-radius: 8px; background-color: #334155; }
QDateEdit::down-arrow { image: url("data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxMCIgaGVpZ2h0PSIxMCIgdmlld0JveD0iMCAwIDEwIDEwIj48cGF0aCBkPSJNMSAzIEg5IEw1IDggWiIgZmlsbD0iI2Y4ZmFmYyIvPjwvc3ZnPg=="); width: 12px; height: 12px; }
QCalendarWidget QWidget { background-color: #1e293b; color: #f8fafc; font-family: 'Malgun Gothic'; }
QCalendarWidget QAbstractItemView:enabled { background-color: #1e293b; color: #f8fafc; selection-background-color: #6366f1; selection-color: white; }
QCalendarWidget QToolButton { color: #f8fafc; background-color: #334155; border-radius: 4px; font-weight: bold; }
QCalendarWidget QMenu { background-color: #1e293b; color: #f8fafc; }
QCalendarWidget QSpinBox { color: #f8fafc; background-color: #1e293b; selection-background-color: #6366f1; }
QPushButton#PrimaryBtn { background-color: #3b82f6; color: white; border-radius: 8px; padding: 10px 20px; font-weight: bold; font-size: 19px; min-height: 40px; min-width: 100px; font-family: 'Malgun Gothic'; }
QPushButton#PrimaryBtn:hover { background-color: #2563eb; }
QPushButton#SecondaryBtn { background-color: #64748b; color: #f8fafc; border-radius: 8px; padding: 10px 20px; font-weight: bold; font-size: 19px; min-height: 40px; min-width: 100px; font-family: 'Malgun Gothic'; }
QPushButton#SecondaryBtn:hover { background-color: #475569; }
QPushButton#DangerBtn { background-color: #ef4444; color: white; border-radius: 8px; padding: 10px 20px; font-weight: bold; font-size: 19px; min-height: 40px; min-width: 100px; font-family: 'Malgun Gothic'; }
QPushButton#DangerBtn:hover { background-color: #dc2626; }
/* 설정 탭 입력 패널: 추가/수정/삭제 한 줄 */
QPushButton#SettingsActPrimary { background-color: #3b82f6; color: white; border-radius: 8px; padding: 8px 10px; font-weight: bold; font-size: 17px; min-height: 40px; min-width: 52px; font-family: 'Malgun Gothic'; }
QPushButton#SettingsActPrimary:hover { background-color: #2563eb; }
QPushButton#SettingsActSecondary { background-color: #64748b; color: #f8fafc; border-radius: 8px; padding: 8px 10px; font-weight: bold; font-size: 17px; min-height: 40px; min-width: 52px; font-family: 'Malgun Gothic'; }
QPushButton#SettingsActSecondary:hover { background-color: #475569; }
QPushButton#SettingsActDanger { background-color: #ef4444; color: white; border-radius: 8px; padding: 8px 10px; font-weight: bold; font-size: 17px; min-height: 40px; min-width: 52px; font-family: 'Malgun Gothic'; }
QPushButton#SettingsActDanger:hover { background-color: #dc2626; }
QComboBox { background-color: #1e293b; border: 1px solid #475569; border-radius: 8px; color: #f8fafc; padding: 5px 15px; font-size: 21px; height: 40px; font-weight: bold; font-family: 'Malgun Gothic'; }
QComboBox QAbstractItemView { background-color: #1e293b; color: #f8fafc; selection-background-color: #3b82f6; border: 1px solid #334155; outline: none; }
QComboBox QAbstractItemView::item { min-height: 35px; padding: 2px 10px; }
QLabel#InputLabel { color: #ffffff; font-weight: bold; font-family: 'Malgun Gothic'; font-size: 18px; }
"""


def _settings_cm_to_px(cm: float) -> int:
    """DPI(논리 인치) 기준 cm → 픽셀 (설정 탭 테이블 너비 보조)."""
    app = QApplication.instance()
    scr = app.primaryScreen() if app else None
    dpi = float(scr.logicalDotsPerInchX()) if scr else 96.0
    return max(1, int(cm * dpi / 2.54))


def _auth_url():
    """관리자 로그인 API URL (API_BASE_URL에서 /api/admin → /api/auth/verify_device_admin)."""
    base = (API_BASE_URL or "").rstrip("/")
    if "/api/admin" in base:
        return base.replace("/api/admin", "/api/auth/verify_device_admin")
    return base.rstrip("/api") + "/api/auth/verify_device_admin"

class AdminLoginDialog(QDialog):
    """식당관리 PC 앱 로그인. 성공 시 token 반환."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.token = None
        self.user = None
        self.setWindowTitle("식당관리 로그인")
        self.setMinimumSize(400, 280)
        self.setStyleSheet(QSS)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(15)
        form = QFormLayout()
        self.emp_no_input = QLineEdit()
        self.emp_no_input.setPlaceholderText("사번")
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("성명")
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("비밀번호 (최초/재설정 시 입력)")
        self.password_input.setEchoMode(QLineEdit.Password)
        form.addRow(QLabel("사번"), self.emp_no_input)
        form.addRow(QLabel("성명"), self.name_input)
        form.addRow(QLabel("비밀번호"), self.password_input)
        layout.addLayout(form)
        btn_layout = QHBoxLayout()
        self.login_btn = QPushButton("로그인")
        self.login_btn.setObjectName("PrimaryBtn")
        self.login_btn.clicked.connect(self.do_login)
        self.cancel_btn = QPushButton("취소")
        self.cancel_btn.setObjectName("SecondaryBtn")
        self.cancel_btn.clicked.connect(self.reject)
        btn_layout.addStretch()
        btn_layout.addWidget(self.login_btn)
        btn_layout.addWidget(self.cancel_btn)
        layout.addLayout(btn_layout)

    def do_login(self):
        emp_no = self.emp_no_input.text().strip()
        name = self.name_input.text().strip()
        password = (self.password_input.text() or "").strip()
        if not emp_no or not name:
            QMessageBox.warning(self, "입력 오류", "사번과 성명을 입력하세요.")
            return
        self.login_btn.setEnabled(False)
        try:
            r = httpx.post(
                _auth_url(),
                json={"emp_no": emp_no, "name": name, "password": password},
                timeout=API_TIMEOUT,
            )
            body = r.json() if r.status_code in (200, 400) else {}
            if r.status_code == 200:
                self.token = body.get("access_token")
                self.user = body.get("user") or {}
                if self.token:
                    self.accept()
                else:
                    QMessageBox.warning(self, "오류", "토큰을 받지 못했습니다.")
            else:
                QMessageBox.warning(self, "로그인 실패", (body.get("detail") or r.text or "로그인에 실패했습니다."))
        except Exception as e:
            QMessageBox.critical(self, "연결 오류", f"서버에 연결할 수 없습니다:\n{e}")
        finally:
            self.login_btn.setEnabled(True)

class EmployeeSearchDialog(QDialog):
    def __init__(self, api, main_win, parent=None):
        super().__init__(parent)
        self.api = api
        self.main_win = main_win
        self.selected_employee = None
        self.setWindowTitle("사원 검색")
        self.setMinimumSize(600, 500)
        self.setStyleSheet(QSS)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Search area
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("검색할 성명 입력...")
        self.search_input.returnPressed.connect(self.on_search)
        
        self.search_btn = QPushButton("조회")
        self.search_btn.setObjectName("PrimaryBtn")
        self.search_btn.setFixedWidth(80)
        self.search_btn.clicked.connect(self.on_search)
        
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.search_btn)
        layout.addLayout(search_layout)
        
        # Table area
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["사번", "성명", "부서", "회사"])
        setup_standard_table(self.table)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.doubleClicked.connect(self.on_select)
        layout.addWidget(self.table)
        
        # Bottom buttons
        btn_layout = QHBoxLayout()
        self.select_btn = QPushButton("선택")
        self.select_btn.setObjectName("PrimaryBtn")
        self.select_btn.clicked.connect(self.on_select)
        
        self.cancel_btn = QPushButton("취소")
        self.cancel_btn.setObjectName("SecondaryBtn")
        self.cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.select_btn)
        btn_layout.addWidget(self.cancel_btn)
        layout.addLayout(btn_layout)

    def on_search(self):
        name = self.search_input.text().strip()
        if not name:
            return
        self.search_btn.setEnabled(False)
        self.loader = DataLoader(self.api.get_employees, name)
        self.loader.finished.connect(self.display_results)
        self.loader.error.connect(self.on_search_error)
        self.loader.start()

    def on_search_error(self, err_msg):
        self.search_btn.setEnabled(True)
        QMessageBox.warning(self, "조회 실패", f"사원 검색 중 오류가 발생했습니다:\n{err_msg}")

    def display_results(self, data):
        self.search_btn.setEnabled(True)
        if not isinstance(data, list):
            return
        self.results = data
        self.table.setRowCount(len(data))
        
        comp_map = {c["id"]: c["name"] for c in self.main_win.companies_data if isinstance(c, dict)}
        dept_map = {d["id"]: d["name"] for d in self.main_win.departments_data if isinstance(d, dict)}
        
        for i, row in enumerate(data):
            comp_name = comp_map.get(row.get("company_id"), "Unknown")
            dept_name = dept_map.get(row.get("department_id"), "Unknown")
            
            item_no = QTableWidgetItem(str(row.get("emp_no", "")))
            item_no.setData(Qt.UserRole, i) # Original index in self.results
            
            self.table.setItem(i, 0, item_no)
            self.table.setItem(i, 1, QTableWidgetItem(str(row.get("name", ""))))
            self.table.setItem(i, 2, QTableWidgetItem(dept_name))
            self.table.setItem(i, 3, QTableWidgetItem(comp_name))
            
    def on_select(self):
        selected = self.table.selectedItems()
        if selected:
            row = selected[0].row()
            item = self.table.item(row, 0)
            if item is not None:
                original_idx = item.data(Qt.UserRole)
                if original_idx is not None and 0 <= original_idx < len(self.results):
                    self.selected_employee = self.results[original_idx]
                    self.accept()
                    return
            QMessageBox.warning(self, "경고", "선택한 행의 데이터를 찾을 수 없습니다.")
        else:
            QMessageBox.warning(self, "경고", "사원을 선택해 주세요.")

def setup_standard_table(table):
    table.verticalHeader().setVisible(False)
    table.verticalHeader().setDefaultSectionSize(38)
    table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
    table.setAlternatingRowColors(True)
    table.setSelectionBehavior(QAbstractItemView.SelectRows)
    table.setSelectionMode(QAbstractItemView.SingleSelection)
    table.setEditTriggers(QAbstractItemView.NoEditTriggers)
    table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    # QSS로 셀·헤더 테두리를 그림. setShowGrid(True)이면 선이 겹쳐 보일 수 있음
    table.setShowGrid(False)
    table.setSortingEnabled(True)

class APIClient:
    def __init__(self, base_url=None, token=None):
        self.base_url = base_url or API_BASE_URL
        self.token = token
        self.client = httpx.Client(timeout=API_TIMEOUT)

    def _auth_headers(self):
        if self.token:
            return {"Authorization": f"Bearer {self.token}"}
        return {}

    def get_stats(self):
        try:
            r = self.client.get(f"{self.base_url}/stats/today", headers=self._auth_headers())
            return r.json() if r.status_code == 200 else None
        except Exception:
            return None


    def get_raw_data(self, search="", start_date=None, end_date=None):
        try:
            params = {"search": search} if search else {}
            if start_date:
                params["start_date"] = start_date
            if end_date:
                params["end_date"] = end_date
            r = self.client.get(f"{self.base_url}/raw-data", params=params, headers=self._auth_headers())
            body = None
            try:
                body = r.json()
            except Exception:
                body = {}
            if r.status_code == 200:
                return (True, body if isinstance(body, list) else [])
            msg = (body or {}).get("detail", r.text or "조회 실패")
            return (False, msg if isinstance(msg, str) else str(msg))
        except Exception as e:
            return (False, str(e))

    def create_manual_raw_data(self, data):
        try:
            r = self.client.post(f"{self.base_url}/raw-data/manual", params=data, headers=self._auth_headers())
            body = None
            try:
                body = r.json()
            except Exception:
                body = {}
            if r.status_code == 200:
                return (True, body)
            detail = (body or {}).get("detail", r.text or "등록 실패")
            return (False, detail)
        except Exception as e:
            return (False, str(e))

    def update_raw_data(self, log_id, data):
        try:
            r = self.client.put(f"{self.base_url}/raw-data/{log_id}", json=data, headers=self._auth_headers())
            body = None
            try:
                body = r.json()
            except Exception:
                body = {}
            if r.status_code == 200:
                return (True, body)
            detail = (body or {}).get("detail", r.text or "수정 실패")
            return (False, detail)
        except Exception as e:
            return (False, str(e))

    def delete_raw_data(self, log_id):
        try:
            r = self.client.delete(f"{self.base_url}/raw-data/{log_id}", headers=self._auth_headers())
            return r.status_code == 200
        except Exception:
            return False

    def void_log(self, log_id, reason):
        try:
            r = self.client.patch(f"{self.base_url}/raw-data/{log_id}/void", json={"reason": reason}, headers=self._auth_headers())
            return r.status_code == 200
        except Exception:
            return False

    # Company Actions
    def get_companies(self):
        try:
            r = self.client.get(f"{self.base_url}/companies", headers=self._auth_headers())
            return r.json() if r.status_code == 200 else []
        except Exception:
            return []

    def create_company(self, code, name):
        try:
            r = self.client.post(f"{self.base_url}/companies", json={"code": code, "name": name}, headers=self._auth_headers())
            body = None
            try:
                body = r.json()
            except Exception:
                body = {}
            if r.status_code == 200:
                return True, body
            detail = (body or {}).get("detail", r.text or "알 수 없는 오류")
            return False, detail
        except Exception as e:
            return False, str(e)

    def update_company(self, cid, code, name):
        try:
            r = self.client.patch(f"{self.base_url}/companies/{cid}", json={"code": code, "name": name}, headers=self._auth_headers())
            return r.json() if r.status_code == 200 else None
        except Exception:
            return None

    def delete_company(self, cid):
        try:
            r = self.client.delete(f"{self.base_url}/companies/{cid}", headers=self._auth_headers())
            return r.status_code == 200
        except Exception:
            return False

    # Policy Actions
    def get_policies(self):
        """(성공 여부, 목록 또는 오류 메시지). 서버 500 등 시 빈 리스트만 주면 UI에 원인이 안 보였음."""
        try:
            r = self.client.get(f"{self.base_url}/policies", headers=self._auth_headers())
            try:
                body = r.json()
            except Exception:
                body = {}
            if r.status_code == 200:
                return (True, _normalize_api_list_payload(body))
            detail = body.get("detail", r.text) if isinstance(body, dict) else r.text
            if isinstance(detail, list):
                detail = str(detail)
            return (False, detail or "조회 실패")
        except Exception as e:
            return (False, str(e))

    def create_policy(self, data):
        try:
            r = self.client.post(f"{self.base_url}/policies", json=data, headers=self._auth_headers())
            if r.status_code in (200, 201):
                return (True, r.json())
            try:
                detail = r.json().get("detail", "등록 실패")
            except Exception:
                detail = r.text or "등록 실패"
            return (False, detail)
        except Exception as e:
            return (False, str(e))

    def update_policy(self, policy_id, data):
        try:
            r = self.client.put(f"{self.base_url}/policies/{policy_id}", json=data, headers=self._auth_headers())
            if r.status_code == 200:
                return (True, r.json())
            try:
                detail = r.json().get("detail", "수정 실패")
            except Exception:
                detail = r.text or "수정 실패"
            return (False, detail)
        except Exception as e:
            return (False, str(e))

    def delete_policy(self, policy_id):
        try:
            r = self.client.delete(f"{self.base_url}/policies/{policy_id}", headers=self._auth_headers())
            return r.status_code == 200
        except: return False

    # Department Actions
    def get_departments(self, company_id=None):
        try:
            params = {"company_id": company_id} if company_id else {}
            r = self.client.get(f"{self.base_url}/departments", params=params, headers=self._auth_headers())
            return r.json() if r.status_code == 200 else []
        except Exception:
            return []

    def create_department(self, company_id, code, name):
        try:
            r = self.client.post(f"{self.base_url}/departments", json={"company_id": company_id, "code": code, "name": name}, headers=self._auth_headers())
            if r.status_code == 200:
                return True, r.json()
            else:
                detail = r.json().get("detail", "알 수 없는 오류")
                return False, detail
        except Exception as e:
            return False, str(e)

    def update_department(self, did, code, name):
        try:
            r = self.client.patch(f"{self.base_url}/departments/{did}", json={"code": code, "name": name}, headers=self._auth_headers())
            return r.json() if r.status_code == 200 else None
        except Exception:
            return None

    def delete_department(self, did):
        try:
            r = self.client.delete(f"{self.base_url}/departments/{did}", headers=self._auth_headers())
            return r.status_code == 200
        except Exception:
            return False

    # Employee Actions
    def get_employees(self, search="", status=None):
        try:
            params = {"search": search}
            if status:
                params["status"] = status
            r = self.client.get(f"{self.base_url}/employees", params=params, headers=self._auth_headers())
            return r.json() if r.status_code == 200 else []
        except Exception:
            return []

    def create_employee(self, data):
        try:
            r = self.client.post(f"{self.base_url}/employees", json=data, headers=self._auth_headers())
            return (True, r.json()) if r.status_code == 200 else (False, r.json().get("detail", "등록 실패"))
        except Exception as e: return (False, str(e))

    def update_employee(self, emp_id, data):
        try:
            r = self.client.put(f"{self.base_url}/employees/{emp_id}", json=data, headers=self._auth_headers())
            return (True, r.json()) if r.status_code == 200 else (False, r.json().get("detail", "수정 실패"))
        except Exception as e: return (False, str(e))

    def delete_employee(self, emp_id, permanent=False):
        try:
            # 서버가 bool 쿼리 파라미터를 확실히 인식하도록 1/0 사용
            params = {"permanent": "1"} if permanent else {}
            r = self.client.delete(f"{self.base_url}/employees/{emp_id}", params=params, headers=self._auth_headers())
            return r.status_code == 200
        except Exception:
            return False

    def get_excel_report_data(self, year, month):
        try:
            r = self.client.get(f"{self.base_url}/reports/excel", params={"year": year, "month": month}, headers=self._auth_headers())
            return r.content if r.status_code == 200 else None
        except Exception:
            return None

    def import_employees_excel(self, company_id, file_content):
        try:
            r = self.client.post(
                f"{self.base_url}/employees/import",
                params={"company_id": company_id},
                content=file_content,
                headers=self._auth_headers()
            )
            return (True, r.json()) if r.status_code == 200 else (False, r.json().get("detail", "가져오기 실패"))
        except Exception as e: return (False, str(e))

    def reset_device_auth(self, emp_id):
        try:
            r = self.client.post(f"{self.base_url}/employees/{emp_id}/reset-device", headers=self._auth_headers())
            return (True, r.json()) if r.status_code == 200 else (False, r.json().get("detail", "초기화 실패"))
        except Exception as e: return (False, str(e))

    def get_notice(self):
        """공지 내용 조회 (백엔드 API)."""
        try:
            r = self.client.get(f"{self.base_url}/notice", headers=self._auth_headers())
            if r.status_code == 200:
                data = r.json()
                return data.get("content", "") or ""
            return ""
        except Exception:
            return ""

    def save_notice_api(self, content: str):
        """공지 내용 저장 (백엔드 API). content는 <br> 포함해 PWA에서 줄바꿈 표시."""
        try:
            r = self.client.put(f"{self.base_url}/notice", json={"content": content}, headers=self._auth_headers())
            return r.status_code == 200
        except Exception:
            return False

    def get_admins(self):
        try:
            r = self.client.get(f"{self.base_url}/admins", headers=self._auth_headers())
            return r.json() if r.status_code == 200 else []
        except Exception:
            return []

    def create_admin(self, emp_no: str, name: str):
        try:
            r = self.client.post(f"{self.base_url}/admins", json={"emp_no": emp_no, "name": name}, headers=self._auth_headers())
            if r.status_code == 200:
                return (True, r.json())
            return (False, (r.json() or {}).get("detail", "등록 실패"))
        except Exception as e:
            return (False, str(e))

    def update_admin(self, user_id: int, name: str):
        try:
            r = self.client.put(f"{self.base_url}/admins/{user_id}", json={"name": name}, headers=self._auth_headers())
            return (r.status_code == 200, (r.json() or {}).get("detail", "수정 실패") if r.status_code != 200 else None)
        except Exception as e:
            return (False, str(e))

    def delete_admin(self, user_id: int):
        try:
            r = self.client.delete(f"{self.base_url}/admins/{user_id}", headers=self._auth_headers())
            return r.status_code == 200
        except Exception:
            return False

    def reset_admin_device(self, user_id: int):
        try:
            r = self.client.post(f"{self.base_url}/admins/{user_id}/reset-device", headers=self._auth_headers())
            return r.status_code == 200
        except Exception:
            return False

    def get_device_settings(self):
        """장치 설정(프린터·경광등) 조회."""
        try:
            r = self.client.get(f"{self.base_url}/settings/device", headers=self._auth_headers())
            return r.json() if r.status_code == 200 else None
        except Exception:
            return None

    def put_device_settings(self, data: dict):
        """장치 설정 저장. 성공 시 (True, 응답), 실패 시 (False, 메시지)."""
        try:
            r = self.client.put(f"{self.base_url}/settings/device", json=data, headers=self._auth_headers())
            body = r.json() if r.status_code in (200, 400) else {}
            if r.status_code == 200:
                return (True, body)
            return (False, (body.get("detail") or r.text or "저장 실패"))
        except Exception as e:
            return (False, str(e))

    def list_qr_terminals(self):
        try:
            r = self.client.get(f"{self.base_url}/terminals", headers=self._auth_headers())
            return r.json() if r.status_code == 200 else []
        except Exception:
            return []

    def get_qr_terminal(self, terminal_id: int):
        try:
            r = self.client.get(f"{self.base_url}/terminals/{terminal_id}", headers=self._auth_headers())
            return r.json() if r.status_code == 200 else None
        except Exception:
            return None

    def create_qr_terminal(self, data: dict):
        try:
            r = self.client.post(f"{self.base_url}/terminals", json=data, headers=self._auth_headers())
            if r.status_code == 200:
                return (True, r.json())
            body = r.json() if r.text else {}
            return (False, body.get("detail", "등록 실패"))
        except Exception as e:
            return (False, str(e))

    def update_qr_terminal(self, terminal_id: int, data: dict):
        try:
            r = self.client.put(f"{self.base_url}/terminals/{terminal_id}", json=data, headers=self._auth_headers())
            if r.status_code == 200:
                return (True, r.json())
            body = r.json() if r.text else {}
            return (False, body.get("detail", "수정 실패"))
        except Exception as e:
            return (False, str(e))

    def delete_qr_terminal(self, terminal_id: int):
        try:
            r = self.client.delete(f"{self.base_url}/terminals/{terminal_id}", headers=self._auth_headers())
            return r.status_code == 200
        except Exception:
            return False

    def close(self):
        self.client.close()

class StatCard(QFrame):
    def __init__(self, title, icon_text, color):
        super().__init__()
        self.setObjectName("StatCard")
        layout = QVBoxLayout(self)
        self.val_label = QLabel("0")
        self.val_label.setObjectName("StatValue")
        self.title_label = QLabel(title)
        self.title_label.setObjectName("StatLabel")
        layout.addWidget(self.val_label)
        layout.addWidget(self.title_label)
    def set_value(self, val):
        self.val_label.setText(str(val))

class DashboardScreen(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        header_h_layout = QHBoxLayout()
        header = QLabel("오늘의 현황")
        header.setObjectName("HeaderTitle")
        
        # Digital Clock
        self.clock_label = QLabel()
        self.clock_label.setStyleSheet("color: #6366f1; font-size: 26px; font-weight: bold; font-family: 'Malgun Gothic'; margin-left: 20px; margin-top: 10px;")
        
        header_h_layout.addWidget(header)
        header_h_layout.addWidget(self.clock_label)
        header_h_layout.addStretch()
        layout.addLayout(header_h_layout)

        # Update clock every second
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_clock)
        self.timer.start(1000)
        self.update_clock()
        
        # Stats layout (Dynamic cards)
        self.stats_layout = QHBoxLayout()
        layout.addLayout(self.stats_layout)
        
        self.cards = {} # Store cards by title
        
        info_frame = QFrame()
        info_frame.setObjectName("StatCard")
        info_layout = QVBoxLayout(info_frame)
        info_layout.setContentsMargins(0, 0, 0, 0) # 테이블이 프레임에 꽉 차도록 여백 제거
        self.recent_table = QTableWidget(0, 4)
        self.recent_table.setHorizontalHeaderLabels(["시간", "이름", "부서", "식사종류"])
        setup_standard_table(self.recent_table) # 표준 테이블 스타일 적용
        
        # Fixed height for exactly 17 rows + header
        # Header(50px) + 17 rows * 38px + border/padding
        self.recent_table.setFixedHeight(700)
        
        info_layout.addWidget(self.recent_table)
        layout.addWidget(info_frame, 1)


    def update_clock(self):
        # 대시보드 시계: 한국 로컬타임(KST) 표시
        kst = timezone(timedelta(hours=9))
        now = datetime.now(kst)
        days = ["월", "화", "수", "목", "금", "토", "일"]
        day_str = days[now.weekday()]
        time_str = now.strftime(f"%m월 %d일({day_str}) %H:%M:%S")
        self.clock_label.setText(time_str)

    def update_stats(self, stats):
        if not stats or not isinstance(stats, dict): return
        
        # Define card data to build
        # Order: Total -> Meal Summaries (Policies) -> Exception
        card_data = []
        card_data.append({"title": "오늘 총 식수", "val": stats.get("total_count", 0), "color": "#6366f1", "icon": "👥"})
        
        # Add meal summaries (Policies)
        colors = ["#8b5cf6", "#10b981", "#f59e0b", "#06b6d4"] # Shared palette
        for i, s in enumerate(stats.get("meal_summaries", [])):
            color = colors[i % len(colors)]
            card_data.append({"title": s["meal_type"], "val": s["count"], "color": color, "icon": "🍱"})
            
        # Add exception (Unified)
        card_data.append({"title": "예외", "val": stats.get("exception_count", 0), "color": "#ef4444", "icon": "⚠️"})
        
        # Refresh cards
        # If card counts match, just update values (for performance)
        if len(self.cards) == len(card_data):
            for item in card_data:
                title = item["title"]
                if title in self.cards:
                    self.cards[title].set_value(item["val"])
                else: 
                    # Rebuild if titles mismatch
                    self._rebuild_cards(card_data)
                    break
        else:
            self._rebuild_cards(card_data)

    def _rebuild_cards(self, card_data):
        # Clear existing
        while self.stats_layout.count():
            child = self.stats_layout.takeAt(0)
            if child.widget(): child.widget().deleteLater()
        self.cards = {}
        
        # Create new
        for item in card_data:
            card = StatCard(item["title"], item["icon"], item["color"])
            card.set_value(item["val"])
            self.stats_layout.addWidget(card)
            self.cards[item["title"]] = card

    def update_recent(self, data):
        if not isinstance(data, list): return
        self.recent_table.setSortingEnabled(False)
        self.recent_table.setUpdatesEnabled(False)
        self.recent_table.setRowCount(len(data))
        kst = timezone(timedelta(hours=9))
        utc = timezone.utc
        for i, row in enumerate(data):
            if not isinstance(row, dict): continue
            user = row.get("user") or {}
            policy = row.get("policy") or {}
            ts = row.get("created_at", "")
            # 서버 created_at: KST 기준 문자열(또는 타임존 있으면 KST로 변환) → 대시보드에는 한국 로컬타임으로 표시
            display_ts = ""
            try:
                if isinstance(ts, str):
                    s = ts.replace("Z", "+00:00").strip().replace(" ", "T")
                    dt = datetime.fromisoformat(s)
                    if dt.tzinfo is None:
                        # 타임존 없음 = 서버가 KST(naive)로 준 값 → 변환 없이 시간만 사용
                        display_ts = dt.strftime("%H:%M")
                    else:
                        display_ts = dt.astimezone(kst).strftime("%H:%M")
                elif hasattr(ts, "astimezone"):
                    dt = ts if ts.tzinfo else ts.replace(tzinfo=utc)
                    display_ts = dt.astimezone(kst).strftime("%H:%M")
                else:
                    display_ts = str(ts)[:8] if ts else ""
            except Exception:
                try:
                    if isinstance(ts, str) and "T" in ts:
                        display_ts = ts.split("T")[-1][:5]
                    elif isinstance(ts, str) and " " in ts:
                        display_ts = ts.split(" ")[-1][:5]
                    else:
                        display_ts = str(ts)[:5] if ts else ""
                except Exception:
                    display_ts = str(ts)[:5] if ts else ""
            self.recent_table.setItem(i, 0, QTableWidgetItem(display_ts[:5] if len(display_ts) >= 5 else display_ts))
            self.recent_table.setItem(i, 1, QTableWidgetItem(str(user.get("name", ""))))
            self.recent_table.setItem(i, 2, QTableWidgetItem(str(user.get("department_name", ""))))
            self.recent_table.setItem(i, 3, QTableWidgetItem(str(policy.get("meal_type", ""))))
        self.recent_table.setUpdatesEnabled(True)
        self.recent_table.setSortingEnabled(True)

class CompanyScreen(QWidget):
    def __init__(self, api, main_win):
        super().__init__()
        self.api = api
        self.main_win = main_win
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        
        header_layout = QHBoxLayout()
        header = QLabel("회사 관리")
        header.setObjectName("HeaderTitle")
        
        reg_frame = QFrame()
        reg_frame.setObjectName("StatCard")
        reg_layout = QHBoxLayout(reg_frame)
        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("회사코드")
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("회사명")
        self.add_btn = QPushButton("등록")
        self.add_btn.setObjectName("PrimaryBtn")
        self.add_btn.setFixedWidth(100)
        self.add_btn.clicked.connect(self.on_add)
        
        self.edit_btn = QPushButton("수정")
        self.edit_btn.setObjectName("SecondaryBtn")
        self.edit_btn.setFixedWidth(100)
        self.edit_btn.clicked.connect(self.on_edit)
        
        self.del_btn = QPushButton("삭제")
        self.del_btn.setObjectName("DangerBtn")
        self.del_btn.setFixedWidth(100)
        self.del_btn.clicked.connect(self.on_delete)
        
        reg_layout.addWidget(self.code_input)
        reg_layout.addWidget(self.name_input)
        reg_layout.addWidget(self.add_btn)
        reg_layout.addWidget(self.edit_btn)
        reg_layout.addWidget(self.del_btn)
        
        header_layout.addWidget(header)
        header_layout.addStretch()
        header_layout.addWidget(reg_frame)
        layout.addLayout(header_layout)
        
        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["회사코드", "회사명"])
        setup_standard_table(self.table)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        layout.addWidget(self.table)

    def on_selection_changed(self):
        selected = self.table.selectedItems()
        if selected:
            row = selected[0].row()
            item0 = self.table.item(row, 0)
            item1 = self.table.item(row, 1)
            self.code_input.setText(item0.text() if item0 else "")
            self.name_input.setText(item1.text() if item1 else "")
            cid = item0.data(Qt.UserRole) if item0 else None
            self.edit_btn.setEnabled(cid is not None)
            self.del_btn.setEnabled(cid is not None)
            self.name_input.setFocus()
            self.name_input.selectAll()
        else:
            self.code_input.clear()
            self.name_input.clear()
            self.edit_btn.setEnabled(False)
            self.del_btn.setEnabled(False)
        
    def load_data(self):
        QApplication.setOverrideCursor(Qt.WaitCursor)
        self.loader = DataLoader(self.api.get_companies)
        self.loader.finished.connect(self.display_data)
        self.loader.start()

    def display_data(self, data):
        QApplication.restoreOverrideCursor()
        if not isinstance(data, list): return
        self.table.setSortingEnabled(False)
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(len(data))
        for i, row in enumerate(data):
            if not isinstance(row, dict): continue
            code = row.get("code", "")
            name = row.get("name", "")
            item_code = QTableWidgetItem(str(code))
            item_code.setData(Qt.UserRole, row.get("id"))
            self.table.setItem(i, 0, item_code)
            self.table.setItem(i, 1, QTableWidgetItem(str(name)))
        self.table.setUpdatesEnabled(True)
        self.table.setSortingEnabled(True)

    def clear_inputs(self):
        self.code_input.clear()
        self.name_input.clear()
        self.table.clearSelection()
        self.edit_btn.setEnabled(False)
        self.del_btn.setEnabled(False)

    def on_edit(self):
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "경고", "수정할 회사를 선택하세요.")
            return

        row = selected[0].row()
        item0 = self.table.item(row, 0)
        cid = item0.data(Qt.UserRole) if item0 else None
        if cid is None:
            QMessageBox.warning(self, "경고", "선택한 항목을 수정할 수 없습니다.")
            return
        code = self.code_input.text()
        name = self.name_input.text()
        
        self.edit_btn.setEnabled(False)
        self.main_win.statusBar().showMessage("수정 중...", 2000)
        self.edit_loader = DataLoader(self.api.update_company, cid, code, name)
        self.edit_loader.finished.connect(self.on_edit_finished)
        self.edit_loader.error.connect(self.on_edit_error)
        self.edit_loader.start()

    def on_edit_error(self, err_msg):
        self.edit_btn.setEnabled(True)
        QMessageBox.warning(self, "오류", f"수정 중 오류가 발생했습니다:\n{err_msg}")

    def on_edit_finished(self, data):
        self.edit_btn.setEnabled(True)
        if data:
            self.main_win.statusBar().showMessage("수정되었습니다.", 3000)
            selected = self.table.selectedItems()
            if selected:
                row = selected[0].row()
                item0 = self.table.item(row, 0)
                if item0 is not None:
                    self.table.setItem(row, 0, QTableWidgetItem(data.get("code", "")))
                    self.table.item(row, 0).setData(Qt.UserRole, data.get("id"))
                self.table.setItem(row, 1, QTableWidgetItem(str(data.get("name", ""))))
            self.main_win.on_company_changed()
            QMessageBox.information(self, "성공", "수정되었습니다.")
            self.clear_inputs()
        else:
            QMessageBox.warning(self, "오류", "수정에 실패했습니다.")

    def on_delete(self):
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "경고", "삭제할 회사를 선택하세요.")
            return

        row = selected[0].row()
        item0 = self.table.item(row, 0)
        item1 = self.table.item(row, 1)
        cid = item0.data(Qt.UserRole) if item0 else None
        if cid is None:
            QMessageBox.warning(self, "경고", "선택한 항목을 삭제할 수 없습니다.")
            return
        name = item1.text() if item1 else "(선택 행)"
        
        reply = QMessageBox.question(self, "삭제 확인", f"'{name}' 회사를 정말 삭제하시겠습니까?\n관련 부서 및 사원 데이터에 영향을 줄 수 있습니다.",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.del_btn.setEnabled(False)
            self.main_win.statusBar().showMessage("삭제 중...", 2000)
            self.del_loader = DataLoader(self.api.delete_company, cid)
            self.del_loader.finished.connect(self.on_delete_finished)
            self.del_loader.error.connect(self.on_del_error)
            self.del_loader.start()

    def on_del_error(self, err_msg):
        self.del_btn.setEnabled(True)
        QMessageBox.warning(self, "오류", f"삭제 중 오류가 발생했습니다:\n{err_msg}")

    def on_delete_finished(self, success):
        self.del_btn.setEnabled(True)
        if success:
            self.main_win.statusBar().showMessage("삭제되었습니다.", 3000)
            selected = self.table.selectedItems()
            if selected:
                row = selected[0].row()
                self.table.removeRow(row) # UI에서 즉시 제거
                self.main_win.on_company_changed()
            QMessageBox.information(self, "성공", "삭제되었습니다.")
            self.clear_inputs()
            self.load_data() # 데이터 서버와 동기화
        else:
            QMessageBox.warning(self, "오류", "삭제에 실패했습니다.")
            
    def on_add(self):
        code = self.code_input.text()
        name = self.name_input.text()
        if not code or not name:
            QMessageBox.warning(self, "경고", "코드와 이름을 입력하세요.")
            return
        
        # Optimistic UI update: Insert row immediately
        pos = self.table.rowCount()
        self.table.insertRow(pos)
        item_code = QTableWidgetItem(code)
        item_code.setData(Qt.UserRole, None) # Pending ID
        self.table.setItem(pos, 0, item_code)
        self.table.setItem(pos, 1, QTableWidgetItem(name))
        
        self.main_win.statusBar().showMessage(f"'{name}' 등록 중...", 2000)
        self.add_btn.setEnabled(False)
        self.add_loader = DataLoader(self.api.create_company, code, name)
        self.add_loader.finished.connect(self.on_add_finished)
        self.add_loader.error.connect(self.on_add_error)
        self.add_loader.start()

    def on_add_error(self, e):
        self.add_btn.setEnabled(True)
        self.load_data() # Rollback optimistic change by reloading
        QMessageBox.critical(self, "치명적 오류", f"스레드 오류: {e}")

    def on_add_finished(self, result):
        self.add_btn.setEnabled(True)
        success, data = result if isinstance(result, tuple) else (False, str(result))
        if success and isinstance(data, dict):
            self.main_win.statusBar().showMessage("회사가 등록되었습니다.", 3000)
            for i in range(self.table.rowCount()):
                item0 = self.table.item(i, 0)
                if item0 is not None and item0.data(Qt.UserRole) is None:
                    item0.setData(Qt.UserRole, data.get("id"))
                    break
            self.main_win.on_company_changed()
            QMessageBox.information(self, "성공", "등록되었습니다.")
            self.clear_inputs()
        else:
            self.load_data()
            msg = data if isinstance(data, str) else "등록 실패"
            QMessageBox.warning(self, "오류", f"등록에 실패했습니다: {msg}")

class DepartmentScreen(QWidget):
    def __init__(self, api, main_win):
        super().__init__()
        self.api = api
        self.main_win = main_win
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        
        header_layout = QHBoxLayout()
        header = QLabel("부서 관리")
        header.setObjectName("HeaderTitle")
        
        reg_frame = QFrame()
        reg_frame.setObjectName("StatCard")
        reg_layout = QHBoxLayout(reg_frame)
        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("부서코드")
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("부서명")
        self.add_btn = QPushButton("등록")
        self.add_btn.setObjectName("PrimaryBtn")
        self.add_btn.setFixedWidth(100)
        self.add_btn.clicked.connect(self.on_add)
        
        self.edit_btn = QPushButton("수정")
        self.edit_btn.setObjectName("SecondaryBtn")
        self.edit_btn.setFixedWidth(100)
        self.edit_btn.clicked.connect(self.on_edit)
        
        self.del_btn = QPushButton("삭제")
        self.del_btn.setObjectName("DangerBtn")
        self.del_btn.setFixedWidth(100)
        self.del_btn.clicked.connect(self.on_delete)
        
        reg_layout.addWidget(self.code_input)
        reg_layout.addWidget(self.name_input)
        reg_layout.addWidget(self.add_btn)
        reg_layout.addWidget(self.edit_btn)
        reg_layout.addWidget(self.del_btn)
        
        header_layout.addWidget(header)
        header_layout.addStretch()
        header_layout.addWidget(reg_frame)
        layout.addLayout(header_layout)
        
        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["부서코드", "부서명"])
        setup_standard_table(self.table)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        layout.addWidget(self.table)

    def on_selection_changed(self):
        selected = self.table.selectedItems()
        if selected:
            row = selected[0].row()
            item0 = self.table.item(row, 0)
            item1 = self.table.item(row, 1)
            dept_id = item0.data(Qt.UserRole) if item0 else None
            self.code_input.setText(item0.text() if item0 else "")
            self.name_input.setText(item1.text() if item1 else "")
            self.edit_btn.setEnabled(dept_id is not None)
            self.del_btn.setEnabled(dept_id is not None)
            self.name_input.setFocus()
            self.name_input.selectAll()
        else:
            self.code_input.clear()
            self.name_input.clear()
            self.edit_btn.setEnabled(False)
            self.del_btn.setEnabled(False)
    def clear_inputs(self):
        self.code_input.clear()
        self.name_input.clear()
        self.table.clearSelection()
        self.edit_btn.setEnabled(False)
        self.del_btn.setEnabled(False)

    def update_company_combo(self, companies):
        pass  # 부서 관리: 회사 1개만 사용, 콤보 없음

    def load_data(self):
        companies = getattr(self.main_win, "companies_data", []) or []
        cid = companies[0]["id"] if companies and isinstance(companies[0], dict) else None
        if cid is None:
            self.table.setRowCount(0)
            return

        try:
            cid_int = int(cid)
        except (TypeError, ValueError):
            cid_int = None
        if cid_int is None:
            self.table.setRowCount(0)
            return

        dept_list = [
            d for d in self.main_win.departments_data
            if isinstance(d, dict) and d.get("company_id") is not None and int(d.get("company_id")) == cid_int
        ]
        self.display_data(dept_list)

    def display_data(self, data):
        QApplication.restoreOverrideCursor()
        if not isinstance(data, list): return
        self.table.setSortingEnabled(False)
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(len(data))

        for i, row in enumerate(data):
            if not isinstance(row, dict): continue
            item_code = QTableWidgetItem(str(row.get("code") or ""))
            item_code.setData(Qt.UserRole, row.get("id"))
            self.table.setItem(i, 0, item_code)
            self.table.setItem(i, 1, QTableWidgetItem(str(row.get("name") or "")))
        self.table.setUpdatesEnabled(True)
        self.table.setSortingEnabled(True)

    def on_edit(self):
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "경고", "수정할 부서를 선택하세요.")
            return
        
        row = selected[0].row()
        did = self.table.item(row, 0).data(Qt.UserRole)
        if did is None:
            QMessageBox.warning(self, "경고", "선택한 항목을 수정할 수 없습니다.")
            return
        code = self.code_input.text()
        name = self.name_input.text()
        
        self.edit_btn.setEnabled(False)
        self.main_win.statusBar().showMessage("부서 수정 중...", 2000)
        self.edit_loader = DataLoader(self.api.update_department, did, code, name)
        self.edit_loader.finished.connect(self.on_edit_finished)
        self.edit_loader.error.connect(self.on_edit_error)
        self.edit_loader.start()

    def on_edit_error(self, err_msg):
        self.edit_btn.setEnabled(True)
        QMessageBox.warning(self, "오류", f"수정 중 오류가 발생했습니다:\n{err_msg}")

    def on_edit_finished(self, data):
        self.edit_btn.setEnabled(True)
        if data:
            self.main_win.statusBar().showMessage("수정되었습니다.", 3000)
            # Update local cache in main window
            for i, d in enumerate(self.main_win.departments_data):
                if d["id"] == data["id"]:
                    self.main_win.departments_data[i] = data
                    break
            
            QMessageBox.information(self, "성공", "수정되었습니다.")
            self.clear_inputs()
            self.load_data()
            self.main_win.employees.update_dept_combo()
            self.main_win.reports._refresh_dept_sub_combo()
        else:
            QMessageBox.warning(self, "오류", "수정에 실패했습니다.")

    def on_delete(self):
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "경고", "삭제할 부서를 선택하세요.")
            return
            
        row = selected[0].row()
        did = self.table.item(row, 0).data(Qt.UserRole)
        if did is None:
            QMessageBox.warning(self, "경고", "선택한 항목을 삭제할 수 없습니다.")
            return
        name_item = self.table.item(row, 1)
        name = name_item.text() if name_item else "(선택 행)"
        
        reply = QMessageBox.question(self, "삭제 확인", f"'{name}' 부서를 정말 삭제하시겠습니까?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.del_btn.setEnabled(False)
            self.main_win.statusBar().showMessage("부서 삭제 중...", 2000)
            self.del_loader = DataLoader(self.api.delete_department, did)
            self.del_loader.finished.connect(self.on_delete_finished)
            self.del_loader.error.connect(self.on_del_error)
            self.del_loader.start()

    def on_del_error(self, err_msg):
        self.del_btn.setEnabled(True)
        QMessageBox.warning(self, "오류", f"삭제 중 오류가 발생했습니다:\n{err_msg}")

    def on_delete_finished(self, success):
        self.del_btn.setEnabled(True)
        if success:
            self.main_win.statusBar().showMessage("삭제되었습니다.", 3000)
            selected = self.table.selectedItems()
            if selected:
                row = selected[0].row()
                target_did = self.table.item(row, 0).data(Qt.UserRole)
                self.main_win.departments_data = [d for d in self.main_win.departments_data if d["id"] != target_did]
                self.table.removeRow(row)
            self.main_win.employees.update_dept_combo()
            self.main_win.reports._refresh_dept_sub_combo()
            self.main_win.sync_departments()
            QMessageBox.information(self, "성공", "삭제되었습니다.")
            self.clear_inputs()
            self.load_data()
        else:
            QMessageBox.warning(self, "오류", "삭제에 실패했습니다.")

    def on_add(self):
        companies = getattr(self.main_win, "companies_data", []) or []
        cid = companies[0]["id"] if companies and isinstance(companies[0], dict) else None
        code = self.code_input.text()
        name = self.name_input.text()
        if not cid or not code or not name:
            QMessageBox.warning(self, "경고", "코드와 부서명을 입력하세요.")
            return

        # Optimistic UI update
        pos = self.table.rowCount()
        self.table.insertRow(pos)
        item_code = QTableWidgetItem(code)
        item_code.setData(Qt.UserRole, None)
        self.table.setItem(pos, 0, item_code)
        self.table.setItem(pos, 1, QTableWidgetItem(name))

        self.add_btn.setEnabled(False)
        self.add_loader = DataLoader(self.api.create_department, cid, code, name)
        self.add_loader.finished.connect(self.on_add_finished)
        self.add_loader.error.connect(self.on_add_error)
        self.add_loader.start()

    def on_add_error(self, e):
        self.add_btn.setEnabled(True)
        self.load_data()
        QMessageBox.critical(self, "치명적 오류", f"스레드 오류: {e}")

    def on_add_finished(self, result):
        self.add_btn.setEnabled(True)
        success, data = result if isinstance(result, tuple) else (False, str(result))
        if success and isinstance(data, dict):
            self.main_win.statusBar().showMessage("부서가 등록되었습니다.", 3000)
            self.main_win.departments_data.append(data)
            self.load_data()
            self.main_win.employees.update_dept_combo()
            self.main_win.reports._refresh_dept_sub_combo()
            QMessageBox.information(self, "성공", "등록되었습니다.")
            self.clear_inputs()
        else:
            self.main_win.sync_departments()
            msg = data if isinstance(data, str) else "등록 실패"
            QMessageBox.warning(self, "오류", f"등록에 실패했습니다: {msg}")

class EmployeeScreen(QWidget):
    def __init__(self, api, main_win):
        super().__init__()
        self.api = api
        self.main_win = main_win
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        header_layout = QHBoxLayout()
        header = QLabel("사원 관리")
        header.setObjectName("HeaderTitle")
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("사번 또는 이름 검색...")
        self.search_input.setFixedWidth(250)
        
        self.refresh_btn = QPushButton("새로고침")
        self.refresh_btn.setFixedWidth(80)
        self.refresh_btn.clicked.connect(self.load_data)
        
        header_layout.addWidget(header)
        header_layout.addStretch()
        header_layout.addWidget(self.refresh_btn)
        header_layout.addWidget(self.search_input)
        layout.addLayout(header_layout)

        # Registration Frame
        reg_frame = QFrame()
        reg_frame.setObjectName("StatCard")
        reg_layout = QHBoxLayout(reg_frame)
        
        self.dept_combo = QComboBox()
        self.dept_combo.setItemDelegate(QStyledItemDelegate())
        self.dept_combo.setPlaceholderText("부서")
        self.dept_combo.setFixedWidth(180)
        
        self.emp_no_input = QLineEdit()
        self.emp_no_input.setPlaceholderText("사번")
        self.emp_no_input.setFixedWidth(120)
        
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("이름")
        self.name_input.setFixedWidth(150)
        
        
        self.add_btn = QPushButton("등록")
        self.add_btn.setObjectName("PrimaryBtn")
        self.add_btn.setFixedWidth(100)
        self.add_btn.clicked.connect(self.on_add)
        
        self.edit_btn = QPushButton("수정")
        self.edit_btn.setObjectName("SecondaryBtn")
        self.edit_btn.setFixedWidth(100)
        self.edit_btn.setEnabled(False)
        self.edit_btn.clicked.connect(self.on_edit)
        
        self.del_btn = QPushButton("삭제(퇴사)")
        self.del_btn.setObjectName("DangerBtn")
        self.del_btn.setFixedWidth(100)
        self.del_btn.setEnabled(False)
        self.del_btn.clicked.connect(self.on_delete)

        self.cancel_resign_btn = QPushButton("퇴사취소")
        self.cancel_resign_btn.setObjectName("SecondaryBtn")
        self.cancel_resign_btn.setFixedWidth(90)
        self.cancel_resign_btn.setEnabled(False)
        self.cancel_resign_btn.setToolTip("선택한 퇴사자를 재직으로 복구합니다.")
        self.cancel_resign_btn.clicked.connect(self.on_cancel_resign)

        self.permanent_del_btn = QPushButton("완전 삭제")
        self.permanent_del_btn.setObjectName("DangerBtn")
        self.permanent_del_btn.setFixedWidth(100)
        self.permanent_del_btn.setEnabled(False)
        self.permanent_del_btn.setToolTip("DB에서 제거하여 같은 사번으로 재등록 가능")
        self.permanent_del_btn.clicked.connect(self.on_permanent_delete)
        
        self.reset_btn = QPushButton("기기 초기화")
        self.reset_btn.setObjectName("SecondaryBtn")
        self.reset_btn.setFixedWidth(110)
        self.reset_btn.setEnabled(False)
        self.reset_btn.clicked.connect(self.on_reset_device)

        self.import_btn = QPushButton("엑셀 일괄등록")
        self.import_btn.setObjectName("SecondaryBtn")
        self.import_btn.setFixedWidth(120)
        self.import_btn.clicked.connect(self.on_import_excel)
        
        reg_layout.addWidget(self.dept_combo)
        reg_layout.addWidget(self.emp_no_input)
        reg_layout.addWidget(self.name_input)
        reg_layout.addWidget(self.add_btn)
        reg_layout.addWidget(self.edit_btn)
        reg_layout.addWidget(self.del_btn)
        reg_layout.addWidget(self.cancel_resign_btn)
        reg_layout.addWidget(self.permanent_del_btn)
        reg_layout.addWidget(self.reset_btn)
        reg_layout.addWidget(self.import_btn)
        layout.addWidget(reg_frame)

        # Table (재직+퇴사 한 테이블, 상태 컬럼 추가)
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["사번", "이름", "부서", "인증", "상태"])
        setup_standard_table(self.table)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        layout.addWidget(self.table)

        # Connections
        self.search_input.textChanged.connect(self.load_data)

    def update_company_combo(self, companies):
        """회사 1개만 사용: 부서 콤보만 갱신."""
        self.update_dept_combo()

    def update_dept_combo(self):
        companies = getattr(self.main_win, "companies_data", []) or []
        cid = companies[0]["id"] if companies and isinstance(companies[0], dict) else None
        self.dept_combo.clear()
        self.dept_combo.addItem("** 선택 **", None)
        if cid is not None:
            try:
                cid_int = int(cid)
                depts = [d for d in self.main_win.departments_data if isinstance(d, dict) and d.get("company_id") is not None and int(d.get("company_id")) == cid_int]
            except (TypeError, ValueError):
                depts = []
            for d in depts:
                self.dept_combo.addItem(d["name"], d["id"])

    def on_selection_changed(self):
        self.edit_btn.setEnabled(False)
        self.del_btn.setEnabled(False)
        self.cancel_resign_btn.setEnabled(False)
        self.permanent_del_btn.setEnabled(False)
        self.reset_btn.setEnabled(False)
        selected = self.table.selectedItems()
        if selected:
            row = selected[0].row()
            item0 = self.table.item(row, 0)
            item1 = self.table.item(row, 1)
            item2 = self.table.item(row, 2)
            item4 = self.table.item(row, 4)
            emp_no = item0.text() if item0 else ""
            name = item1.text() if item1 else ""
            dept_name = item2.text() if item2 else ""
            status_text = (item4.text() if item4 else "").strip()

            self.emp_no_input.setText(emp_no)
            self.name_input.setText(name)

            idx = self.dept_combo.findText(dept_name)
            if idx >= 0:
                self.dept_combo.setCurrentIndex(idx)

            if item0 and item0.data(Qt.UserRole) is not None:
                self.edit_btn.setEnabled(True)
                self.permanent_del_btn.setEnabled(True)
                if status_text == "재직":
                    self.del_btn.setEnabled(True)
                    self.reset_btn.setEnabled(True)
                    self.cancel_resign_btn.setEnabled(False)
                else:
                    self.del_btn.setEnabled(False)
                    self.reset_btn.setEnabled(False)
                    self.cancel_resign_btn.setEnabled(True)

    def clear_inputs(self):
        self.emp_no_input.clear()
        self.name_input.clear()
        self.table.clearSelection()
        self.edit_btn.setEnabled(False)
        self.del_btn.setEnabled(False)
        self.cancel_resign_btn.setEnabled(False)
        self.permanent_del_btn.setEnabled(False)
        self.reset_btn.setEnabled(False)

    def on_add(self):
        companies = getattr(self.main_win, "companies_data", []) or []
        cid = companies[0]["id"] if companies and isinstance(companies[0], dict) else None
        did = self.dept_combo.currentData()
        emp_no = self.emp_no_input.text().strip()
        name = self.name_input.text().strip()
        
        if not cid or not emp_no or not name:
            QMessageBox.warning(self, "경고", "사번과 이름을 입력하세요.")
            return

        payload = {
            "company_id": cid,
            "department_id": did,
            "emp_no": emp_no,
            "name": name,
            "status": "ACTIVE"
        }
        
        self.main_win.statusBar().showMessage("사원 등록 중...", 5000)
        self.add_btn.setEnabled(False)
        self.loader = DataLoader(self.api.create_employee, payload)
        self.loader.finished.connect(self.on_action_finished)
        self.loader.error.connect(self.on_action_error)
        self.loader.start()

    def on_edit(self):
        selected = self.table.selectedItems()
        if not selected: return
        row = selected[0].row()
        item0 = self.table.item(row, 0)
        eid = item0.data(Qt.UserRole) if item0 else None
        if eid is None: return
        
        companies = getattr(self.main_win, "companies_data", []) or []
        cid = companies[0]["id"] if companies and isinstance(companies[0], dict) else None
        did = self.dept_combo.currentData()
        emp_no = self.emp_no_input.text().strip()
        name = self.name_input.text().strip()

        payload = {
            "company_id": cid,
            "department_id": did,
            "emp_no": emp_no,
            "name": name
        }
        
        self.loader = DataLoader(self.api.update_employee, eid, payload)
        self.loader.finished.connect(self.on_action_finished)
        self.loader.start()

    def on_import_excel(self):
        companies = getattr(self.main_win, "companies_data", []) or []
        cid = companies[0]["id"] if companies and isinstance(companies[0], dict) else None
        if not cid:
            QMessageBox.warning(self, "경고", "회사 데이터가 없습니다.")
            return
            
        from PyQt5.QtWidgets import QFileDialog
        file_path, _ = QFileDialog.getOpenFileName(self, "사원 엑셀 파일 선택", "", "Excel Files (*.xlsx *.xls)")
        if not file_path: return
        
        try:
            with open(file_path, "rb") as f:
                file_content = f.read()
            
            self.import_btn.setEnabled(False)
            self.main_win.statusBar().showMessage("엑셀 데이터 업로드 중...", 5000)
            self.import_loader = DataLoader(self.api.import_employees_excel, cid, file_content)
            self.import_loader.finished.connect(self.on_import_finished)
            self.import_loader.start()
        except Exception as e:
            QMessageBox.critical(self, "오류", f" 파일을 읽을 수 없습니다: {str(e)}")

    def on_import_finished(self, result):
        self.import_btn.setEnabled(True)
        success, data = result if isinstance(result, tuple) else (False, result)
        if success:
            msg = data.get("message", "완료되었습니다.")
            QMessageBox.information(self, "임포트 성공", msg)
            self.load_data() # Refresh table
        else:
            QMessageBox.warning(self, "임포트 실패", str(data))

    def on_delete(self):
        selected = self.table.selectedItems()
        if not selected: return
        row = selected[0].row()
        item0 = self.table.item(row, 0)
        eid = item0.data(Qt.UserRole) if item0 else None
        if eid is None: return
        
        if QMessageBox.question(self, "확인", "정말 이 사원을 삭제(퇴사) 처리하시겠습니까?\n(DB에는 남아 있어 같은 사번으로 재등록 시 '재등록' 처리됩니다.)") == QMessageBox.Yes:
            self.del_btn.setEnabled(False)
            self.permanent_del_btn.setEnabled(False)
            self.loader = DataLoader(self.api.delete_employee, eid, False)
            self.loader.finished.connect(self.on_delete_finished)
            self.loader.error.connect(self.on_action_error)
            self.loader.start()

    def on_cancel_resign(self):
        selected = self.table.selectedItems()
        if not selected:
            return
        row = selected[0].row()
        item0 = self.table.item(row, 0)
        item4 = self.table.item(row, 4)
        eid = item0.data(Qt.UserRole) if item0 else None
        if eid is None:
            return
        status_text = (item4.text() if item4 else "").strip()
        if status_text != "퇴사":
            QMessageBox.warning(self, "경고", "퇴사자만 퇴사취소할 수 있습니다.")
            return
        name_item = self.table.item(row, 1)
        name = name_item.text() if name_item else "(선택 행)"
        if QMessageBox.question(self, "확인", f"'{name}' 사원을 재직으로 복구하시겠습니까?") != QMessageBox.Yes:
            return
        self.cancel_resign_btn.setEnabled(False)
        self.loader = DataLoader(self.api.update_employee, eid, {"status": "ACTIVE"})
        self.loader.finished.connect(self.on_action_finished)
        self.loader.error.connect(self.on_action_error)
        self.loader.start()

    def on_permanent_delete(self):
        selected = self.table.selectedItems()
        if not selected: return
        row = selected[0].row()
        item0 = self.table.item(row, 0)
        item4 = self.table.item(row, 4)
        eid = item0.data(Qt.UserRole) if item0 else None
        if eid is None: return
        name_item = self.table.item(row, 1)
        name = name_item.text() if name_item else "(선택 행)"
        status_text = (item4.text() if item4 else "").strip()

        if status_text == "재직":
            msg = (
                f"'{name}' 사원은 재직자입니다.\n"
                "완전 삭제하면 DB에서 제거되며, 모든 식사 기록도 삭제됩니다.\n"
                "정말 삭제하시겠습니까?"
            )
        else:
            msg = (
                f"'{name}' 사원을 DB에서 완전히 제거합니다.\n"
                "같은 사번으로 다시 등록할 수 있습니다.\n"
                "관련 식사 기록도 삭제됩니다. 계속하시겠습니까?"
            )
        if QMessageBox.question(
            self, "완전 삭제 확인", msg,
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        ) == QMessageBox.Yes:
            self.del_btn.setEnabled(False)
            self.permanent_del_btn.setEnabled(False)
            self.loader = DataLoader(self.api.delete_employee, eid, True)
            self.loader.finished.connect(self.on_delete_finished)
            self.loader.error.connect(self.on_action_error)
            self.loader.start()

    def on_delete_finished(self, success):
        self.del_btn.setEnabled(True)
        self.permanent_del_btn.setEnabled(True)
        if success:
            # 서버 상태 기준으로 다시 로드해 삭제 반영 보장
            self.clear_inputs()
            self.load_data()
            QMessageBox.information(self, "성공", "삭제되었습니다.")
        else:
            QMessageBox.warning(self, "오류", "삭제에 실패했습니다.")

    def on_reset_device(self):
        selected = self.table.selectedItems()
        if not selected: return
        row = selected[0].row()
        item0 = self.table.item(row, 0)
        eid = item0.data(Qt.UserRole) if item0 else None
        if eid is None: return
        name_item = self.table.item(row, 1)
        name = name_item.text() if name_item else "(선택 행)"
        
        if QMessageBox.question(self, "확인", f"'{name}' 사원의 기기 인증 상태 및 비밀번호를 초기화하시겠습니까?") == QMessageBox.Yes:
            self.reset_btn.setEnabled(False)
            self.loader = DataLoader(self.api.reset_device_auth, eid)
            self.loader.finished.connect(self.on_action_finished)
            self.loader.error.connect(self.on_action_error)
            self.loader.start()

    def on_action_finished(self, result):
        self.add_btn.setEnabled(True)
        self.edit_btn.setEnabled(True)
        self.reset_btn.setEnabled(True)
        self.cancel_resign_btn.setEnabled(True)
        if isinstance(result, tuple):
            success, data = result
        else:
            success, data = bool(result), None
        if success:
            QMessageBox.information(self, "성공", "작업이 완료되었습니다.")
            self.load_data()
            self.clear_inputs()
        else:
            msg = str(data) if data is not None else "작업에 실패했습니다."
            QMessageBox.warning(self, "오류", msg)

    def on_action_error(self, err_msg):
        self.add_btn.setEnabled(True)
        self.edit_btn.setEnabled(True)
        self.del_btn.setEnabled(True)
        self.cancel_resign_btn.setEnabled(True)
        self.permanent_del_btn.setEnabled(True)
        self.reset_btn.setEnabled(True)
        QMessageBox.critical(self, "네트워크 오류", f"서버 통신 중 오류가 발생했습니다:\n{err_msg}")

    def load_data(self):
        # 재직+퇴사 모두 조회 (status 없음 = 전체)
        self.loader = DataLoader(self.api.get_employees, self.search_input.text(), None)
        self.loader.finished.connect(self.display_data)
        self.loader.start()

    def display_data(self, data):
        if not isinstance(data, list): return

        selected_ids = set()
        for item in self.table.selectedItems():
            if item.column() == 0 and item.data(Qt.UserRole) is not None:
                selected_ids.add(item.data(Qt.UserRole))

        companies = getattr(self.main_win, "companies_data", []) or []
        cid = companies[0]["id"] if companies and isinstance(companies[0], dict) else None
        if cid is not None:
            try:
                cid_int = int(cid)
                data = [r for r in data if isinstance(r, dict) and r.get("company_id") is not None and int(r.get("company_id")) == cid_int]
            except (TypeError, ValueError):
                pass

        self.table.setSortingEnabled(False)
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(len(data))
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["사번", "이름", "부서", "인증", "상태"])

        dept_map = {d["id"]: d["name"] for d in self.main_win.departments_data if isinstance(d, dict)}

        for i, row in enumerate(data):
            if not isinstance(row, dict): continue
            item_no = QTableWidgetItem(str(row.get("emp_no", "")))
            row_id = row.get("id")
            item_no.setData(Qt.UserRole, row_id)

            dept_name = dept_map.get(row.get("department_id"), "Unknown")
            status_label = "재직" if row.get("status") == "ACTIVE" else "퇴사"

            self.table.setItem(i, 0, item_no)
            self.table.setItem(i, 1, QTableWidgetItem(str(row.get("name", ""))))
            self.table.setItem(i, 2, QTableWidgetItem(dept_name))

            is_verified = row.get("is_verified", False)
            auth_status = "O" if is_verified else "X"
            item_auth = QTableWidgetItem(auth_status)
            if not is_verified:
                item_auth.setForeground(QColor("#ef4444"))
            else:
                item_auth.setForeground(QColor("#10b981"))
            self.table.setItem(i, 3, item_auth)

            item_status = QTableWidgetItem(status_label)
            self.table.setItem(i, 4, item_status)

            if row.get("status") == "RESIGNED":
                for col in range(5):
                    it = self.table.item(i, col)
                    if it: it.setForeground(QColor("#999999"))

            if row_id in selected_ids:
                for col in range(5):
                    it = self.table.item(i, col)
                    if it: it.setSelected(True)

        self.table.setUpdatesEnabled(True)
        self.table.setSortingEnabled(True)

class RawDataScreen(QWidget):
    def __init__(self, api, main_win):
        super().__init__()
        self.api = api
        self.main_win = main_win
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        
        header_layout = QHBoxLayout()
        header = QLabel("원시 데이터 관리")
        header.setObjectName("HeaderTitle")
        
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(QDate.currentDate())
        self.start_date_edit.setFixedWidth(180)

        tilde = QLabel("~")
        tilde.setStyleSheet("color: white; font-size: 18px; font-weight: bold;")
        
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setFixedWidth(180)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("사번 또는 성명 검색...")
        self.search_input.setFixedWidth(250)
        self.search_input.returnPressed.connect(self.load_data)

        self.search_btn = QPushButton("조회")
        self.search_btn.setObjectName("PrimaryBtn")
        self.search_btn.setFixedWidth(80)
        self.search_btn.clicked.connect(self.load_data)
        
        header_layout.addWidget(header)
        header_layout.addStretch()
        header_layout.addWidget(self.start_date_edit)
        header_layout.addWidget(tilde)
        header_layout.addWidget(self.end_date_edit)
        header_layout.addSpacing(10)
        header_layout.addWidget(self.search_input)
        header_layout.addWidget(self.search_btn)
        self.raw_excel_btn = QPushButton("엑셀")
        self.raw_excel_btn.setObjectName("SecondaryBtn")
        self.raw_excel_btn.setFixedWidth(70)
        self.raw_excel_btn.clicked.connect(self.on_raw_excel)
        header_layout.addWidget(self.raw_excel_btn)

        # Main layout structure: Left (Table) / Right (Inputs)
        main_h_layout = QHBoxLayout()
        layout.addLayout(main_h_layout)
        
        # Left side: Table Area
        left_layout = QVBoxLayout()
        left_layout.addLayout(header_layout)
        
        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(["No", "날짜", "시간", "이름", "사번", "식사종류", "경로", "상태"])
        setup_standard_table(self.table)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.table.setColumnWidth(0, 80)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        left_layout.addWidget(self.table)
        main_h_layout.addLayout(left_layout, 7) # 70% width
        
        # Right side: Input Panel
        self.input_panel = QFrame()
        self.input_panel.setObjectName("StatCard")
        self.input_panel.setFixedWidth(350)
        input_layout = QVBoxLayout(self.input_panel)
        input_layout.setSpacing(15)
        
        panel_title = QLabel("기록 등록/수정")
        panel_title.setStyleSheet("font-size: 22px; font-weight: bold; color: #6366f1; margin-bottom: 10px;")
        input_layout.addWidget(panel_title)
        
        # 날짜만 선택 (시간은 식사 종류에 따라 자동 적용)
        date_label = QLabel("날짜")
        date_label.setObjectName("InputLabel")
        self.edit_date = QDateEdit()
        self.edit_date.setCalendarPopup(True)
        self.edit_date.setDate(QDate.currentDate())
        self.edit_date.setFixedHeight(40)
        date_hint = QLabel("※ 시간은 선택한 식사 종류에 따라 자동 적용됩니다.")
        date_hint.setStyleSheet("color: #94a3b8; font-size: 12px; margin-bottom: 4px;")
        
        # Employee Search
        emp_label = QLabel("사원 검색 (이름)")
        emp_label.setObjectName("InputLabel")
        search_h = QHBoxLayout()
        self.emp_search_input = QLineEdit()
        self.emp_search_input.setPlaceholderText("검색 버튼을 클릭하세요...")
        self.emp_search_input.setReadOnly(True)
        self.emp_search_btn = QPushButton("사원검색")
        self.emp_search_btn.setObjectName("PrimaryBtn")
        self.emp_search_btn.setFixedWidth(100)
        self.emp_search_btn.clicked.connect(self.on_search_employee)
        search_h.addWidget(self.emp_search_input)
        search_h.addWidget(self.emp_search_btn)
        
        self.selected_emp_label = QLabel("선택된 사원: 없음")
        self.selected_emp_label.setStyleSheet("color: #94a3b8; font-size: 16px;")
        self.selected_user_id = None
        
        # Meal Policy (Meal Type)
        policy_label = QLabel("식사 종류")
        policy_label.setObjectName("InputLabel")
        self.policy_combo = QComboBox()
        self.policy_combo.setFixedHeight(40)
        self.policy_combo.addItem("선택하세요", None)
        self.policy_combo.currentIndexChanged.connect(self.on_policy_changed)
        
        # Policies data for auto-time
        self.policies_list = []
        self.load_policies()
        
        guest_label = QLabel("게스트 인원")
        guest_label.setObjectName("InputLabel")
        self.edit_guest = QLineEdit("0")
        self.edit_guest.setFixedHeight(40)
        
        # Add to input layout
        for label, widget in [
            (emp_label, None), (None, search_h), (None, self.selected_emp_label),
            (date_label, self.edit_date), (None, date_hint),
            (policy_label, self.policy_combo), (guest_label, self.edit_guest)
        ]:
            if label: input_layout.addWidget(label)
            if widget: 
                if isinstance(widget, QLayout): input_layout.addLayout(widget)
                else: input_layout.addWidget(widget)
        
        # 기록 등록 버튼: 게스트 인원 밑에 배치
        self.add_btn = QPushButton("기록 등록")
        self.add_btn.setObjectName("PrimaryBtn")
        self.add_btn.setFixedHeight(40)
        self.add_btn.clicked.connect(self.on_add)
        input_layout.addWidget(self.add_btn)
        
        # 기록 수정 버튼: 기록 등록과 같은 간격으로 바로 아래
        self.edit_btn = QPushButton("기록 수정")
        self.edit_btn.setObjectName("SecondaryBtn")
        self.edit_btn.setFixedHeight(40)
        self.edit_btn.clicked.connect(self.on_update)
        self.edit_btn.setEnabled(False)
        input_layout.addWidget(self.edit_btn)
        
        # 기록 삭제 버튼: 기록 수정과 같은 간격으로 바로 아래
        self.del_btn = QPushButton("기록 삭제")
        self.del_btn.setObjectName("DangerBtn")
        self.del_btn.setFixedHeight(40)
        self.del_btn.clicked.connect(self.on_delete)
        self.del_btn.setEnabled(False)
        input_layout.addWidget(self.del_btn)
        
        # 입력창 초기화: 기록 삭제와 같은 간격·같은 크기로 바로 아래
        self.clear_btn = QPushButton("입력창 초기화")
        self.clear_btn.setObjectName("SecondaryBtn")
        self.clear_btn.clicked.connect(self.clear_inputs)
        self.clear_btn.setFixedHeight(40)
        input_layout.addWidget(self.clear_btn)
        
        input_layout.addStretch()
        
        main_h_layout.addWidget(self.input_panel)
        self.current_log_id = None
    def load_data(self):
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
        self.loader = DataLoader(self.api.get_raw_data, self.search_input.text(), start_date, end_date)
        self.loader.finished.connect(self.display_data)
        self.loader.start()
    def display_data(self, data):
        if isinstance(data, tuple) and len(data) >= 2:
            success, payload = data[0], data[1]
            if not success:
                QMessageBox.warning(self, "오류", f"데이터를 가져오는데 실패했습니다.\n{payload}")
                return
            data = payload
        if not isinstance(data, list):
            return
        self.table.setSortingEnabled(False)
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(len(data))
        self.full_data = data
        for i, row in enumerate(data):
            if not isinstance(row, dict): continue
            created_at = row.get("created_at", "")
            date_part = created_at.split("T")[0]
            time_part = created_at.split("T")[-1][:8]
            
            # Row Number (No) - Use numeric data for sorting
            item_no = QTableWidgetItem()
            item_no.setData(Qt.DisplayRole, i + 1)
            row_id = row.get("id")
            item_no.setData(Qt.UserRole, row_id)
            
            self.table.setItem(i, 0, item_no)
            self.table.setItem(i, 1, QTableWidgetItem(date_part))
            self.table.setItem(i, 2, QTableWidgetItem(time_part))
            self.table.setItem(i, 3, QTableWidgetItem(str(row.get("user", {}).get("name", ""))))
            self.table.setItem(i, 4, QTableWidgetItem(str(row.get("user", {}).get("emp_no", ""))))
            
            # 식사종류: 서버에서 내려준 해당 로그의 policy.meal_type 사용 (시간 재계산 시 번외로 바뀌는 현상 방지)
            meal_type = str(row.get("policy", {}).get("meal_type", "") or "번외")
            item_type = QTableWidgetItem(meal_type)
            if meal_type == "번외":
                item_type.setForeground(QColor("#94a3b8")) # Gray color for extra
            self.table.setItem(i, 5, item_type)
            
            self.table.setItem(i, 6, QTableWidgetItem(str(row.get("path", ""))))
            status = "취소됨" if row.get("is_void") else "정상"
            self.table.setItem(i, 7, QTableWidgetItem(status))
        self.table.setUpdatesEnabled(True)
        self.table.setSortingEnabled(True)

    def load_policies(self):
        self.pol_loader = DataLoader(self.api.get_policies)
        self.pol_loader.finished.connect(self.on_policies_loaded)
        self.pol_loader.start()

    def on_policies_loaded(self, data):
        if isinstance(data, tuple) and len(data) == 2:
            ok, payload = data
            if not ok:
                self.policies_list = []
                self.policy_combo.clear()
                self.policy_combo.addItem("선택하세요", None)
                return
            data = payload
        if not isinstance(data, list):
            return
        self.policies_list = data
        self.policy_combo.clear()
        self.policy_combo.addItem("선택하세요", None)
        for p in data:
            self.policy_combo.addItem(p["meal_type"], p["id"])

    def on_policy_changed(self, idx):
        pass  # 시간은 등록/수정 시 식사 종류 기준으로 자동 적용

    def _created_at_time_from_policy(self, policy_id):
        """선택한 식사 종류의 시작 시간 + 5분을 HH:mm:ss 로 반환. 없으면 12:00:00."""
        if not policy_id:
            return "12:00:00"
        policy = next((p for p in self.policies_list if p["id"] == policy_id), None)
        if not policy or not policy.get("start_time"):
            return "12:00:00"
        try:
            h, m = map(int, policy["start_time"].split(":")[:2])
            qtime = QTime(h, m).addSecs(300)
            return qtime.toString("HH:mm:ss")
        except Exception:
            return "12:00:00"

    def get_meal_type_by_time(self, time_str):
        if not time_str: return "번외"
        try:
            h, m, s = map(int, time_str.split(":"))
            target_time = QTime(h, m, s)
            
            for p in self.policies_list:
                if not p.get("start_time") or not p.get("end_time"): continue
                
                # Times from API are HH:MM:SS
                sh, sm, ss = map(int, p["start_time"].split(":")[:3])
                eh, em, es = map(int, p["end_time"].split(":")[:3])
                
                start = QTime(sh, sm, ss)
                end = QTime(eh, em, es)
                
                if start <= end:
                    if start <= target_time <= end:
                        return p["meal_type"]
                else: # Overnight
                    if target_time >= start or target_time <= end:
                        return p["meal_type"]
            return "번외"
        except Exception as e:
            print(f"Error judging meal type: {e}")
            return "번외"

    def on_search_employee(self):
        dialog = EmployeeSearchDialog(self.api, self.main_win, self)
        if dialog.exec_():
            emp = dialog.selected_employee
            if emp:
                self.selected_user_id = emp["id"]
                self.emp_search_input.setText(emp['name'])
                self.selected_emp_label.setText(f"선택됨: {emp['name']} ({emp['emp_no']})")

    def on_selection_changed(self):
        selected = self.table.selectedItems()
        if not selected:
            self.current_log_id = None
            self.add_btn.setEnabled(True)
            self.edit_btn.setEnabled(False)
            self.del_btn.setEnabled(False)
            return

        row_idx = selected[0].row()
        item0 = self.table.item(row_idx, 0)
        log_id = item0.data(Qt.UserRole) if item0 else None
        if log_id is None:
            self.current_log_id = None
            return
        # Find log by log_id in full_data
        log = next((x for x in self.full_data if x["id"] == log_id), None)
        if not log: return
        
        self.current_log_id = log["id"]
        
        # Fill inputs
        dt = datetime.fromisoformat(log["created_at"].replace("Z", ""))
        self.edit_date.setDate(QDate(dt.year, dt.month, dt.day))
        
        self.selected_user_id = log["user_id"]
        self.selected_emp_label.setText(f"선택됨: {log['user']['name']} ({log['user']['emp_no']})")
        
        # Select policy in combo
        policy_id = log["policy_id"]
        for i in range(self.policy_combo.count()):
            if self.policy_combo.itemData(i) == policy_id:
                self.policy_combo.setCurrentIndex(i)
                break
        
        self.edit_guest.setText(str(log.get("guest_count", 0)))
        
        self.add_btn.setEnabled(False)
        self.edit_btn.setEnabled(True)
        self.del_btn.setEnabled(True)

    def clear_inputs(self):
        self.edit_date.setDate(QDate.currentDate())
        self.emp_search_input.clear()
        self.selected_user_id = None
        self.selected_emp_label.setText("선택된 사원: 없음")
        self.policy_combo.setCurrentIndex(0)
        self.edit_guest.setText("0")
        self.table.clearSelection()
        self.current_log_id = None
        self.add_btn.setEnabled(True)
        self.edit_btn.setEnabled(False)
        self.del_btn.setEnabled(False)

    def on_raw_excel(self):
        """원시데이터 테이블을 엑셀 파일로 저장 후 열기"""
        import os
        import tempfile
        import subprocess
        import sys
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            QMessageBox.warning(self, "오류", "엑셀 저장을 위해 openpyxl이 필요합니다.\npip install openpyxl")
            return
        start_str = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_str = self.end_date_edit.date().toString("yyyy-MM-dd")
        title = "원시데이터"
        period_str = f"조회기간: {start_str} ~ {end_str}"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "원시데이터"
        col_count = self.table.columnCount()
        if col_count > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=16)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.cell(row=2, column=1, value=period_str)
        for c in range(col_count):
            h = self.table.horizontalHeaderItem(c)
            ws.cell(row=3, column=c + 1, value=h.text() if h else "").font = Font(bold=True)
        for r in range(self.table.rowCount()):
            for c in range(col_count):
                item = self.table.item(r, c)
                val = item.text() if item else ""
                ws.cell(row=r + 4, column=c + 1, value=val)
        for i in range(1, col_count + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
        try:
            fd, save_path = tempfile.mkstemp(suffix=".xlsx", prefix="RawData_")
            os.close(fd)
            wb.save(save_path)
        except Exception as e:
            QMessageBox.warning(self, "오류", f"엑셀 파일 생성 실패:\n{e}")
            return
        try:
            if sys.platform == "win32":
                os.startfile(save_path)
            elif sys.platform == "darwin":
                subprocess.run(["open", save_path], check=True)
            else:
                subprocess.run(["xdg-open", save_path], check=True)
        except Exception:
            pass

    def on_add(self):
        if not self.selected_user_id:
            QMessageBox.warning(self, "경고", "사원을 먼저 선택하세요.")
            return
        policy_id = self.policy_combo.currentData()
        if not policy_id:
            QMessageBox.warning(self, "경고", "식사 종류를 선택하세요.")
            return
        time_str = self._created_at_time_from_policy(policy_id)
        dt_str = f"{self.edit_date.date().toString('yyyy-MM-dd')}T{time_str}"
        data = {
            "user_id": self.selected_user_id,
            "policy_id": policy_id,
            "created_at": dt_str,
            "guest_count": int(self.edit_guest.text() or 0)
        }
        
        self.action_loader = DataLoader(self.api.create_manual_raw_data, data)
        self.action_loader.finished.connect(self.on_action_finished)
        self.action_loader.start()

    def on_update(self):
        if not self.current_log_id: return
        policy_id = self.policy_combo.currentData()
        time_str = self._created_at_time_from_policy(policy_id)
        dt_str = f"{self.edit_date.date().toString('yyyy-MM-dd')}T{time_str}"
        data = {
            "user_id": self.selected_user_id,
            "policy_id": policy_id,
            "created_at": dt_str,
            "guest_count": int(self.edit_guest.text() or 0)
        }
        self.edit_btn.setEnabled(False)
        self.action_loader = DataLoader(self.api.update_raw_data, self.current_log_id, data)
        self.action_loader.finished.connect(self.on_action_finished)
        self.action_loader.start()

    def on_delete(self):
        if not self.current_log_id: return
        if QMessageBox.question(self, "삭제 확인", "정말로 이 기록을 완전히 삭제하시겠습니까?", 
                               QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return
        self.del_btn.setEnabled(False)
        self.action_loader = DataLoader(self.api.delete_raw_data, self.current_log_id)
        self.action_loader.finished.connect(self.on_action_finished)
        self.action_loader.start()

    def on_action_finished(self, result):
        success = False
        message = ""
        if isinstance(result, tuple): # Add/Update
            success, detail = result
            message = "등록/수정되었습니다." if success else str(detail)
        else: # Delete
            success = result
            message = "삭제되었습니다." if success else "삭제 실패"
            
        if success:
            QMessageBox.information(self, "성공", message)
            selected = self.table.selectedItems()
            if selected and "삭제" in message:
                self.table.removeRow(selected[0].row())
            self.load_data()
            self.clear_inputs()
        else:
            QMessageBox.warning(self, "오류", f"작업 실패: {message}")
            self.edit_btn.setEnabled(True)
            self.del_btn.setEnabled(True)
    
    def on_void_header(self):
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "경고", "취소할 기록을 선택하세요.")
            return
        
        row_idx = selected[0].row()
        log_id = self.table.item(row_idx, 0).data(Qt.UserRole)
        if log_id is None:
            QMessageBox.warning(self, "경고", "선택한 행의 기록 ID를 찾을 수 없습니다.")
            return
        log = next((x for x in self.full_data if x.get("id") == log_id), None)
        if log is None:
            QMessageBox.warning(self, "경고", "선택한 기록을 찾을 수 없습니다.")
            return
        if log.get("is_void"):
            QMessageBox.warning(self, "경고", "이미 취소된 기록입니다.")
            return
            
        self.on_void(log_id)
    def on_void(self, log_id):
        reason, ok = QInputDialog.getText(self, "취소 사유", "취소 사유를 입력하세요:")
        if ok and reason:
            if self.api.void_log(log_id, reason):
                QMessageBox.information(self, "성공", "취소 처리가 완료되었습니다.")
                # Local UI update for performance
                selected = self.table.selectedItems()
                if selected:
                    row_idx = selected[0].row()
                    self.full_data[row_idx]["is_void"] = True
                    self.table.setItem(row_idx, 7, QTableWidgetItem("취소됨"))
            else:
                QMessageBox.warning(self, "오류", "처리에 실패했습니다.")

class PolicyScreen(QWidget):
    def __init__(self, api, main_win):
        super().__init__()
        self.api = api
        self.main_win = main_win
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        
        header_layout = QHBoxLayout()
        
        reg_frame = QFrame()
        reg_frame.setObjectName("StatCard")
        reg_layout = QHBoxLayout(reg_frame)
        self.type_input = QLineEdit()
        self.type_input.setPlaceholderText("식사 종류 (예: 조식)")
        self.type_input.setFixedWidth(250)
        self.start_input = QTimeEdit()
        self.start_input.setDisplayFormat("HH:mm:ss")
        self.start_input.setFixedWidth(180)
        self.end_input = QTimeEdit()
        self.end_input.setDisplayFormat("HH:mm:ss")
        self.end_input.setFixedWidth(180)
        self.price_input = QLineEdit()
        self.price_input.setPlaceholderText("단가 (원)")
        self.price_input.setFixedWidth(180)
        
        self.add_btn = QPushButton("등록")
        self.add_btn.setObjectName("PrimaryBtn")
        self.add_btn.setFixedWidth(100)
        self.add_btn.clicked.connect(self.on_add)
        
        self.edit_btn = QPushButton("수정")
        self.edit_btn.setObjectName("SecondaryBtn")
        self.edit_btn.setFixedWidth(100)
        self.edit_btn.setEnabled(False)
        self.edit_btn.clicked.connect(self.on_edit)
        
        self.del_btn = QPushButton("삭제")
        self.del_btn.setObjectName("DangerBtn")
        self.del_btn.setFixedWidth(100)
        self.del_btn.setEnabled(False)
        self.del_btn.clicked.connect(self.on_delete)
        
        reg_layout.addWidget(self.type_input)
        reg_layout.addWidget(self.start_input)
        reg_layout.addWidget(self.end_input)
        reg_layout.addWidget(self.price_input)
        reg_layout.addWidget(self.add_btn)
        reg_layout.addWidget(self.edit_btn)
        reg_layout.addWidget(self.del_btn)
        
        header_layout.addWidget(reg_frame)
        layout.addLayout(header_layout)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["식사 종류", "시작 시간", "종료 시간", "기본 단가"])
        setup_standard_table(self.table)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        layout.addWidget(self.table, 1)
        
        help_label = QLabel("* 시간 형식: HH:MM:SS (예: 12:00:00)")
        help_label.setStyleSheet("color: #94a3b8; font-size: 14px; font-weight: bold; font-family: 'Malgun Gothic';")
        layout.addWidget(help_label)
        layout.addStretch()
        self._policy_load_seq = 0

    def load_data(self):
        self._policy_load_seq += 1
        seq = self._policy_load_seq
        QApplication.setOverrideCursor(Qt.WaitCursor)
        self.loader = DataLoader(self.api.get_policies)
        self.loader.finished.connect(lambda d, s=seq: self._on_policies_list_loaded(d, s))
        self.loader.error.connect(lambda e, s=seq: self._on_policies_list_error(e, s))
        self.loader.start()

    def _on_policies_list_error(self, err_msg, seq):
        if seq != self._policy_load_seq:
            return
        QApplication.restoreOverrideCursor()
        QMessageBox.warning(self, "식사 정책", f"목록을 불러오지 못했습니다.\n{err_msg}")

    def _on_policies_list_loaded(self, data, seq):
        if seq != self._policy_load_seq:
            return
        if isinstance(data, tuple) and len(data) == 2:
            ok, payload = data
            if not ok:
                QApplication.restoreOverrideCursor()
                QMessageBox.warning(
                    self,
                    "식사 정책",
                    "목록을 불러오지 못했습니다.\n"
                    + (str(payload) if payload is not None else "알 수 없는 오류"),
                )
                return
            data = payload
        self.display_data(data)

    def display_data(self, data):
        QApplication.restoreOverrideCursor()
        if not isinstance(data, list):
            data = _normalize_api_list_payload(data)
        if not isinstance(data, list):
            data = []
        self.table.setSortingEnabled(False)
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(len(data))
        meal_names = {"breakfast": "조식", "lunch": "중식", "dinner": "석식"}
        
        for i, row in enumerate(data):
            if not isinstance(row, dict): continue
            meal_type = row.get("meal_type", "")
            row_id = row.get("id")
            item_meal = QTableWidgetItem(meal_type)
            item_meal.setData(Qt.UserRole, row_id)
            item_meal.setData(Qt.UserRole + 1, meal_type)
            self.table.setItem(i, 0, item_meal)
            self.table.setItem(i, 1, QTableWidgetItem(str(row.get("start_time", ""))))
            self.table.setItem(i, 2, QTableWidgetItem(str(row.get("end_time", ""))))

            item_price = QTableWidgetItem(str(row.get("base_price", 0)))
            self.table.setItem(i, 3, item_price)
        self.table.setUpdatesEnabled(True)
        self.table.setSortingEnabled(True)

    def on_selection_changed(self):
        selected = self.table.selectedItems()
        if selected:
            row = selected[0].row()
            item0 = self.table.item(row, 0)
            item1 = self.table.item(row, 1)
            item2 = self.table.item(row, 2)
            item3 = self.table.item(row, 3)
            meal_display = item0.text() if item0 else ""
            start_time = item1.text() if item1 else ""
            end_time = item2.text() if item2 else ""
            price = item3.text() if item3 else ""

            self.type_input.setText(meal_display)

            self.start_input.setTime(QTime.fromString(start_time, "HH:mm:ss"))
            self.end_input.setTime(QTime.fromString(end_time, "HH:mm:ss"))
            self.price_input.setText(price)

            self.edit_btn.setEnabled(item0 is not None and item0.data(Qt.UserRole) is not None)
            self.del_btn.setEnabled(item0 is not None and item0.data(Qt.UserRole) is not None)
        else:
            self.type_input.clear()
            self.start_input.setTime(QTime(0, 0, 0))
            self.end_input.setTime(QTime(0, 0, 0))
            self.price_input.clear()
            self.edit_btn.setEnabled(False)
            self.del_btn.setEnabled(False)

    def clear_inputs(self):
        self.type_input.clear()
        self.start_input.setTime(QTime(0, 0, 0))
        self.end_input.setTime(QTime(0, 0, 0))
        self.price_input.clear()
        self.table.clearSelection()
        self.edit_btn.setEnabled(False)
        self.del_btn.setEnabled(False)

    def on_add(self):
        meal_type = self.type_input.text().strip()
        start_time = self.start_input.time().toString("HH:mm:ss")
        end_time = self.end_input.time().toString("HH:mm:ss")
        price = self.price_input.text().strip()
        
        if not meal_type or not start_time or not end_time or not price:
            QMessageBox.warning(self, "경고", "모든 정보를 입력하세요.")
            return

        try:
            price_val = int(price)
            data = {
                "meal_type": meal_type,
                "start_time": start_time,
                "end_time": end_time,
                "base_price": price_val,
                "guest_price": price_val,
                "is_active": True
            }
            self.add_btn.setEnabled(False)
            self.loader = DataLoader(self.api.create_policy, data)
            self.loader.finished.connect(self.on_action_finished)
            self.loader.start()
        except ValueError:
            QMessageBox.warning(self, "경고", "단가는 숫자여야 합니다.")

    def on_edit(self):
        selected = self.table.selectedItems()
        if not selected: return
        row = selected[0].row()
        item0 = self.table.item(row, 0)
        pid = item0.data(Qt.UserRole) if item0 else None
        if pid is None: return
        
        meal_type = self.type_input.text().strip()
        start_time = self.start_input.time().toString("HH:mm:ss")
        end_time = self.end_input.time().toString("HH:mm:ss")
        price = self.price_input.text().strip()
        
        if not meal_type or not start_time or not end_time or not price:
            QMessageBox.warning(self, "경고", "모든 정보를 입력하세요.")
            return
        
        try:
            price_val = int(price)
            data = {
                "meal_type": meal_type,
                "start_time": start_time,
                "end_time": end_time,
                "base_price": price_val,
                "guest_price": price_val,
                "is_active": True
            }
            self.edit_btn.setEnabled(False)
            self.loader = DataLoader(self.api.update_policy, pid, data)
            self.loader.finished.connect(self.on_action_finished)
            self.loader.start()
        except ValueError:
            QMessageBox.warning(self, "경고", "단가는 숫자여야 합니다.")

    def on_delete(self):
        selected = self.table.selectedItems()
        if not selected: return
        row = selected[0].row()
        item0 = self.table.item(row, 0)
        pid = item0.data(Qt.UserRole) if item0 else None
        if pid is None: return
        meal_name = item0.text() if item0 else "(선택 행)"
        
        if QMessageBox.question(self, "확인", f"'{meal_name}' 정책을 삭제하시겠습니까?") == QMessageBox.Yes:
            self.del_btn.setEnabled(False)
            self.loader = DataLoader(self.api.delete_policy, pid)
            self.loader.finished.connect(self.on_delete_finished)
            self.loader.start()

    def on_action_finished(self, result):
        self.add_btn.setEnabled(True)
        self.edit_btn.setEnabled(True)
        if isinstance(result, tuple):
            success, data = result
        else:
            success, data = bool(result), None
        if success:
            QMessageBox.information(self, "성공", "작업이 완료되었습니다.")
            self.load_data()
            self.clear_inputs()
        else:
            msg = str(data) if data is not None else "작업에 실패했습니다."
            QMessageBox.warning(self, "오류", msg)

    def on_delete_finished(self, success):
        self.del_btn.setEnabled(True)
        if success:
            selected = self.table.selectedItems()
            if selected:
                self.table.removeRow(selected[0].row())
            QMessageBox.information(self, "성공", "삭제되었습니다.")
            self.load_data()
            self.clear_inputs()
        else:
            QMessageBox.warning(self, "오류", "삭제에 실패했습니다.")

class ReportScreen(QWidget):
    def __init__(self, api, main_win):
        super().__init__()
        self.api = api
        self.main_win = main_win
        self.full_data = []
        self.report_detail_employee = None  # 개인별 상세로 선택된 사원
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # Header
        header_layout = QHBoxLayout()
        header = QLabel("식사 통계 보고서")
        header.setObjectName("HeaderTitle")
        header_layout.addWidget(header)
        header_layout.addStretch()
        layout.addLayout(header_layout)
        
        # Controls Frame
        controls_frame = QFrame()
        controls_frame.setObjectName("StatCard")
        controls_layout = QHBoxLayout(controls_frame)
        controls_layout.setSpacing(15)
        
        date_label = QLabel("조회 기간")
        date_label.setObjectName("InputLabel")
        
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        today = QDate.currentDate()
        self.start_date_edit.setDate(QDate(today.year(), today.month(), 1))
        self.start_date_edit.setFixedWidth(180)
        
        tilde = QLabel("~")
        tilde.setStyleSheet("color: white; font-size: 18px; font-weight: bold;")
        
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setFixedWidth(180)
        
        self.dept_sub_combo = QComboBox()
        self.dept_sub_combo.setFixedWidth(180)
        self._refresh_dept_sub_combo()
        self.dept_sub_combo.currentIndexChanged.connect(self._on_dept_combo_changed)
        
        # 개인별 상세: 레이블 + 검색 버튼 (원시데이터 사원조회와 동일한 검색 다이얼로그)
        self.report_detail_label = QLabel("개인별 상세")
        self.report_detail_label.setObjectName("InputLabel")
        self.report_detail_search_btn = QPushButton("사원검색")
        self.report_detail_search_btn.setObjectName("PrimaryBtn")
        self.report_detail_search_btn.setFixedWidth(100)
        self.report_detail_search_btn.clicked.connect(self._on_report_detail_search)
        self.report_detail_name_edit = QLineEdit()
        self.report_detail_name_edit.setPlaceholderText("검색 버튼을 클릭하세요...")
        self.report_detail_name_edit.setReadOnly(True)
        self.report_detail_name_edit.setFixedWidth(160)
        self.report_detail_name_edit.setMinimumHeight(36)
        
        self.download_btn = QPushButton("엑셀")
        self.download_btn.setObjectName("SecondaryBtn")
        self.download_btn.setToolTip("엑셀 다운로드")
        self.download_btn.setFixedWidth(70)
        self.download_btn.clicked.connect(self.on_download_excel)
        
        controls_layout.addWidget(date_label)
        controls_layout.addWidget(self.start_date_edit)
        controls_layout.addWidget(tilde)
        controls_layout.addWidget(self.end_date_edit)
        controls_layout.addSpacing(20)
        controls_layout.addWidget(self.dept_sub_combo)
        controls_layout.addSpacing(15)
        controls_layout.addWidget(self.report_detail_label)
        controls_layout.addWidget(self.report_detail_search_btn)
        controls_layout.addWidget(self.report_detail_name_edit)
        controls_layout.addStretch()
        controls_layout.addWidget(self.download_btn)
        
        layout.addWidget(controls_frame)
        
        # Table Area
        self.table = QTableWidget(0, 6)
        setup_standard_table(self.table)
        layout.addWidget(self.table, 1)  # stretch: 테이블이 남는 공간 차지

        # 합계 표시 (테이블 바로 아래, 항상 맨 마지막에 고정)
        footer_bar = QFrame()
        footer_bar.setObjectName("ReportFooterBar")
        footer_bar.setFixedHeight(40)
        footer_bar.setStyleSheet("QFrame#ReportFooterBar { background: transparent; }")
        footer_layout = QHBoxLayout(footer_bar)
        footer_layout.setContentsMargins(8, 4, 8, 4)
        footer_layout.addStretch()
        self.report_footer_label = QLabel("합계: 0(0)")
        self.report_footer_label.setObjectName("ReportFooterLabel")
        self.report_footer_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.report_footer_label.setStyleSheet("font-weight: bold; padding: 4px 8px; color: #e2e8f0; font-size: 17px; font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;")
        footer_layout.addWidget(self.report_footer_label)
        layout.addWidget(footer_bar)  # 테이블 다음에 추가 → 항상 하단

    def _refresh_dept_sub_combo(self):
        self.dept_sub_combo.clear()
        self.dept_sub_combo.addItem("개인별", "INDIVIDUAL")   # 개인별 조회 → 개인별 집계 테이블
        self.dept_sub_combo.addItem("전체부서", None)         # 전체부서/부서별 → 상세 리스트 테이블
        depts = getattr(self.main_win, "departments_data", []) or []
        seen = set()
        for d in depts:
            if isinstance(d, dict):
                n = d.get("name") or "미지정"
                if n not in seen:
                    seen.add(n)
                    self.dept_sub_combo.addItem(n, d.get("id"))
        self.dept_sub_combo.setCurrentIndex(0)

    def _on_dept_combo_changed(self):
        self.report_detail_employee = None
        self.report_detail_name_edit.clear()
        self.display_data()

    def _on_report_detail_search(self):
        dialog = EmployeeSearchDialog(self.api, self.main_win, self)
        if dialog.exec_():
            emp = dialog.selected_employee
            if emp:
                self.report_detail_employee = emp
                self.report_detail_name_edit.setText(emp.get("name") or "")
                self.display_data()

    def reset_report(self):
        """다른 창 갔다 왔을 때 보고서 화면 초기화 (조회는 하지 않음)"""
        self.full_data = []
        self.report_detail_employee = None
        self.report_detail_name_edit.clear()
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["사번", "성명", "부서", "식사 횟수", "합계 식수", "총 금액"])
        self.report_footer_label.setText("합계: 0(0)")
        self.table.setSortingEnabled(True)

    def load_data(self):
        start = self.start_date_edit.date().toString("yyyy-MM-dd")
        end = self.end_date_edit.date().toString("yyyy-MM-dd")
        
        self.main_win.statusBar().showMessage("데이터 조회 중...", 3000)
        
        # Get raw data for the range
        self.loader = DataLoader(self.api.get_raw_data, "", start, end)
        self.loader.finished.connect(self.on_data_loaded)
        self.loader.start()

    def on_data_loaded(self, data):
        if isinstance(data, tuple) and len(data) >= 2:
            success, payload = data[0], data[1]
            if not success:
                QMessageBox.warning(self, "오류", f"데이터를 가져오는데 실패했습니다.\n{payload}")
                return
            data = payload
        if not isinstance(data, list):
            QMessageBox.warning(self, "오류", "데이터를 가져오는데 실패했습니다.")
            return

        # Filter out voided logs
        self.full_data = [d for d in data if not d.get("is_void")]
        self.display_data()

    def display_data(self):
        # 부서 콤보로 테이블 구분: 개인별 → 집계 테이블, 전체부서/부서별 → 상세 리스트 테이블
        # 개인별 상세(검색으로 선택한 사원)가 있으면 해당 사원 상세만 표시 (컬럼은 전체부서와 동일)
        self.table.setUpdatesEnabled(False)
        self.table.setSortingEnabled(False)
        self.table.clearSelection()

        if self.report_detail_employee:
            emp_id = self.report_detail_employee.get("id")
            data = [log for log in self.full_data if (log.get("user") or {}).get("id") == emp_id]
            self.table.setColumnCount(6)
            self.table.setHorizontalHeaderLabels(["날짜", "시간", "식사명", "부서", "이름", "사번"])
            data = sorted(data, key=lambda log: log.get("created_at") or "")
            total_amount = 0
            total_meals = 0
            for log in data:
                guest = log.get("guest_count") or log.get("guestCount") or 0
                price = log.get("final_price") or log.get("finalPrice") or 0
                try:
                    guest = int(guest)
                except (TypeError, ValueError):
                    guest = 0
                try:
                    price = int(price)
                except (TypeError, ValueError):
                    price = 0
                total_amount += price * (1 + guest)
                total_meals += (1 + guest)
            self.table.setRowCount(len(data))
            for i, log in enumerate(data):
                u = log.get("user") or log.get("User") or {}
                if not isinstance(u, dict):
                    u = {}
                created = log.get("created_at")
                if isinstance(created, dict):
                    date_part = str(created.get("date", created.get("date_time", "")))[:10]
                    time_part = str(created.get("time", ""))[:8]
                elif isinstance(created, str):
                    created = created.strip()
                    if "T" in created:
                        date_part = created.split("T")[0]
                        time_part = (created.split("T")[-1].replace("Z", "").strip())[:8]
                    else:
                        date_part = created[:10] if len(created) >= 10 else created
                        time_part = ""
                else:
                    date_part = ""
                    time_part = ""
                meal_type = str((log.get("policy") or {}).get("meal_type") or "번외")
                dept_name = u.get("department_name")
                if not dept_name and isinstance(u.get("department"), dict):
                    dept_name = (u.get("department") or {}).get("name")
                dept_name = str(dept_name or log.get("department_name") or "")
                name_str = str(u.get("name") or log.get("user_name") or "")
                emp_no_str = str(u.get("emp_no") or u.get("empNo") or log.get("emp_no") or "")
                self.table.setItem(i, 0, QTableWidgetItem(date_part))
                self.table.setItem(i, 1, QTableWidgetItem(time_part))
                self.table.setItem(i, 2, QTableWidgetItem(meal_type))
                self.table.setItem(i, 3, QTableWidgetItem(dept_name))
                self.table.setItem(i, 4, QTableWidgetItem(name_str))
                self.table.setItem(i, 5, QTableWidgetItem(emp_no_str))
            self.report_footer_label.setText(f"합계: {total_amount:,}({total_meals})")
            self.table.setSortingEnabled(True)
            self.table.setUpdatesEnabled(True)
            return

        dept_id = self.dept_sub_combo.currentData()
        data = self.full_data if dept_id in (None, "INDIVIDUAL") else [
            log for log in self.full_data
            if (log.get("user") or {}).get("department_id") == dept_id
        ]

        if dept_id == "INDIVIDUAL":
            # 개인별 조회 → 사번, 성명, 부서, 식사 횟수, 합계 식수, 총 금액
            self.table.setColumnCount(6)
            self.table.setHorizontalHeaderLabels(["사번", "성명", "부서", "식사 횟수", "합계 식수", "총 금액"])
            agg = {}
            for log in data:
                u = log.get("user") or {}
                if not isinstance(u, dict):
                    u = {}
                key = u.get("id")
                if key is None:
                    continue
                if key not in agg:
                    agg[key] = {"cnt": 0, "guest": 0, "total": 0, "amount": 0, "meta": {"no": u.get("emp_no"), "name": u.get("name"), "dept": u.get("department_name")}}
                guest = log.get("guest_count") or log.get("guestCount") or 0
                price = log.get("final_price") or log.get("finalPrice") or 0
                try:
                    guest = int(guest)
                except (TypeError, ValueError):
                    guest = 0
                try:
                    price = int(price)
                except (TypeError, ValueError):
                    price = 0
                agg[key]["cnt"] += 1
                agg[key]["guest"] += guest
                agg[key]["total"] += (1 + guest)
                agg[key]["amount"] += price * (1 + guest)
            sorted_keys = sorted(agg.keys(), key=lambda k: str(agg[k]["meta"].get("name", "")))
            self.table.setRowCount(len(agg))
            total_amount = 0
            total_meals = 0
            for i, key in enumerate(sorted_keys):
                item = agg[key]
                meta = item["meta"]
                total_amount += item["amount"]
                total_meals += item["total"]
                self.table.setItem(i, 0, QTableWidgetItem(str(meta.get("no", ""))))
                self.table.setItem(i, 1, QTableWidgetItem(str(meta.get("name", ""))))
                self.table.setItem(i, 2, QTableWidgetItem(str(meta.get("dept", ""))))
                self.table.setItem(i, 3, QTableWidgetItem(f"{item['cnt']:,}"))
                self.table.setItem(i, 4, QTableWidgetItem(f"{item['total']:,}"))
                amount_item = QTableWidgetItem(f"{item['amount']:,}원")
                amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(i, 5, amount_item)
            self.report_footer_label.setText(f"합계: {total_amount:,}({total_meals})")
        else:
            # 전체부서 / 부서별 조회 → 날짜, 시간, 식사명, 부서, 이름, 사번 (상세 리스트)
            self.table.setColumnCount(6)
            self.table.setHorizontalHeaderLabels(["날짜", "시간", "식사명", "부서", "이름", "사번"])
            data = sorted(data, key=lambda log: log.get("created_at") or "")
            total_amount = 0
            total_meals = 0
            for log in data:
                guest = log.get("guest_count") or log.get("guestCount") or 0
                price = log.get("final_price") or log.get("finalPrice") or 0
                try:
                    guest = int(guest)
                except (TypeError, ValueError):
                    guest = 0
                try:
                    price = int(price)
                except (TypeError, ValueError):
                    price = 0
                total_amount += price * (1 + guest)
                total_meals += (1 + guest)
            self.table.setRowCount(len(data))
            for i, log in enumerate(data):
                u = log.get("user") or log.get("User") or {}
                if not isinstance(u, dict):
                    u = {}
                created = log.get("created_at")
                if isinstance(created, dict):
                    date_part = str(created.get("date", created.get("date_time", "")))[:10]
                    time_part = str(created.get("time", ""))[:8]
                elif isinstance(created, str):
                    created = created.strip()
                    if "T" in created:
                        date_part = created.split("T")[0]
                        time_part = (created.split("T")[-1].replace("Z", "").strip())[:8]
                    else:
                        date_part = created[:10] if len(created) >= 10 else created
                        time_part = ""
                else:
                    date_part = ""
                    time_part = ""
                meal_type = str((log.get("policy") or {}).get("meal_type") or "번외")
                dept_name = u.get("department_name")
                if not dept_name and isinstance(u.get("department"), dict):
                    dept_name = (u.get("department") or {}).get("name")
                dept_name = str(dept_name or log.get("department_name") or "")
                name_str = str(u.get("name") or log.get("user_name") or "")
                emp_no_str = str(u.get("emp_no") or u.get("empNo") or log.get("emp_no") or "")
                self.table.setItem(i, 0, QTableWidgetItem(date_part))
                self.table.setItem(i, 1, QTableWidgetItem(time_part))
                self.table.setItem(i, 2, QTableWidgetItem(meal_type))
                self.table.setItem(i, 3, QTableWidgetItem(dept_name))
                self.table.setItem(i, 4, QTableWidgetItem(name_str))
                self.table.setItem(i, 5, QTableWidgetItem(emp_no_str))
            self.report_footer_label.setText(f"합계: {total_amount:,}({total_meals})")

        self.table.setSortingEnabled(True)
        self.table.setUpdatesEnabled(True)

    def update_meal_type_summary(self):
        # Removed as requested
        pass

    def on_download_excel(self):
        # 현재 보고서 테이블을 엑셀 파일로 생성 후 열기 → 사용자가 보면서 직접 저장
        import os
        import tempfile
        import subprocess
        import sys

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            QMessageBox.warning(self, "오류", "엑셀 저장을 위해 openpyxl이 필요합니다.\npip install openpyxl")
            return

        dept_id = self.dept_sub_combo.currentData()
        if self.report_detail_employee:
            title = "개인별 상세 보고서 - " + (self.report_detail_employee.get("name") or "선택 사원")
        elif dept_id == "INDIVIDUAL":
            title = "개인별 보고서"
        elif dept_id is None:
            title = "전체부서 보고서"
        else:
            title = self.dept_sub_combo.currentText() + " 보고서"

        start_str = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_str = self.end_date_edit.date().toString("yyyy-MM-dd")
        period_str = f"조회기간: {start_str} ~ {end_str}"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "보고서"

        col_count = self.table.columnCount()

        # 1행: 타이틀 (전체 병합, 가운데, 글자 크게)
        if col_count > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        c1 = ws.cell(row=1, column=1, value=title)
        c1.font = Font(bold=True, size=16)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # 2행: 첫 컬럼에 조회기간
        ws.cell(row=2, column=1, value=period_str)

        # 3행: 헤더
        for c in range(col_count):
            h = self.table.horizontalHeaderItem(c)
            ws.cell(row=3, column=c + 1, value=h.text() if h else "")
        header_font = Font(bold=True)
        for c in range(1, col_count + 1):
            ws.cell(row=3, column=c).font = header_font

        # 데이터 행
        for r in range(self.table.rowCount()):
            for c in range(col_count):
                item = self.table.item(r, c)
                val = item.text() if item else ""
                ws.cell(row=r + 4, column=c + 1, value=val)

        # 합계 행 (우측 정렬)
        footer_text = self.report_footer_label.text()
        next_row = self.table.rowCount() + 4
        right_align = Alignment(horizontal="right", vertical="center")

        if dept_id == "INDIVIDUAL" and not self.report_detail_employee:
            fc = ws.cell(row=next_row, column=1, value=footer_text)
            fc.font = Font(bold=True)
            fc.alignment = right_align
            if col_count > 1:
                ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=col_count)
                ws.cell(row=next_row, column=1).alignment = right_align
        else:
            parts = footer_text.split(" ", 1)
            label_part = parts[0] if parts else "합계:"
            value_part = parts[1] if len(parts) > 1 else ""
            for col_idx, val in enumerate([label_part, value_part], start=col_count - 1):
                cell = ws.cell(row=next_row, column=col_idx, value=val)
                cell.font = Font(bold=True)
                cell.alignment = right_align

        # A4 양식에 맞춘 컬럼 너비 (조회기간이 첫 컬럼에 들어가도록)
        if col_count == 6:
            widths = [24, 10, 8, 10, 10, 10]  # 첫 컬럼 24로 조회기간 수용
        else:
            widths = [12] * (col_count - 1) + [24]
        for i, w in enumerate(widths[:col_count], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 임시 파일로 저장 후 엑셀 열기 (저장은 사용자가 엑셀에서 수행)
        try:
            fd, save_path = tempfile.mkstemp(suffix=".xlsx", prefix="MealReport_")
            os.close(fd)
            wb.save(save_path)
        except Exception as e:
            QMessageBox.warning(self, "오류", f"엑셀 파일 생성 실패:\n{e}")
            return

        try:
            if sys.platform == "win32":
                os.startfile(save_path)
            elif sys.platform == "darwin":
                subprocess.run(["open", save_path], check=True)
            else:
                subprocess.run(["xdg-open", save_path], check=True)
        except Exception:
            pass

    def on_download_finished(self, data, save_path):
        """레거시: API 엑셀 다운로드 완료 (현재는 사용 안 함)"""
        self.download_btn.setEnabled(True)
        if data:
            with open(save_path, "wb") as f:
                f.write(data)
            QMessageBox.information(self, "성공", f"Excel 보고서가 저장되었습니다.\n{save_path}")
        else:
            QMessageBox.warning(self, "오류", "Excel 보고서 다운로드에 실패했습니다.")


class NoticeScreen(QWidget):
    """PWA 공지사항 편집 화면. static/notice.html에 저장하면 PWA 홈에서 표시됨."""
    DEFAULT_NOTICE = """• 오늘 점심 메뉴는 제육볶음과 된장국입니다.<br>
• 12시 30분부터 배식 시작합니다. 현장 도착 시 QR 스캔 부탁드립니다.<br>
• 문의사항은 식당 담당자(내선 1234)로 연락 주세요."""

    def __init__(self, main_win=None):
        super().__init__()
        self.main_win = main_win
        self.notice_path = NOTICE_HTML_PATH
        self.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Maximum)
        self.setObjectName("NoticeScreen")
        self.setAutoFillBackground(False)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 0, 24, 10)
        layout.setSpacing(0)

        content_block = QWidget()
        content_block.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Maximum)
        block_layout = QVBoxLayout(content_block)
        block_layout.setContentsMargins(0, 0, 0, 0)
        block_layout.setSpacing(0)

        header = QLabel("PWA 공지사항")
        header.setObjectName("HeaderTitle")
        header.setContentsMargins(0, 0, 0, 0)
        header.setFixedHeight(40)
        block_layout.addWidget(header)
        block_layout.addSpacing(20)
        hint = QLabel("아래 내용이 PWA 홈 화면의 공지사항 영역에 표시됩니다. 줄바꿈은 그대로 반영되며, <br> 입력 시 HTML 줄바꿈으로 표시됩니다.")
        hint.setObjectName("NoticeHint")
        hint.setWordWrap(True)
        hint.setContentsMargins(0, 0, 0, 0)
        hint.setStyleSheet("color: #94a3b8; font-size: 17px; font-weight: bold; font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;")
        block_layout.addWidget(hint)

        block_layout.addSpacing(4)
        self.text_edit = QPlainTextEdit()
        self.text_edit.setPlaceholderText("공지 내용을 입력하세요...")
        self.text_edit.setFixedSize(400, 200)
        self.text_edit.setStyleSheet("""
            QPlainTextEdit { background-color: #1e293b; color: #f1f5f9; border: 1px solid #475569; border-radius: 8px; padding: 12px 12px 2px 12px; font-size: 19px; font-family: 'Malgun Gothic', 'Segoe UI', sans-serif; font-weight: bold; }
        """)
        block_layout.addWidget(self.text_edit, 0, Qt.AlignLeft)
        block_layout.addSpacing(2)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.save_btn = QPushButton("저장")
        self.save_btn.setObjectName("PrimaryBtn")
        self.save_btn.setFixedWidth(120)
        self.save_btn.clicked.connect(self.save_notice)
        btn_row.addWidget(self.save_btn)
        btn_wrapper = QWidget()
        btn_wrapper.setFixedWidth(400)
        btn_wrapper.setLayout(btn_row)
        block_layout.addWidget(btn_wrapper, 0, Qt.AlignLeft)

        layout.addWidget(content_block, 0, Qt.AlignTop)

        self.load_notice()

    def load_notice(self):
        """백엔드 API에서 공지 로드. 저장된 <br>는 편집용으로 \\n으로 표시."""
        api = getattr(self.main_win, "api", None) if self.main_win else None
        if api:
            raw = api.get_notice()
            if raw:
                self.text_edit.setPlainText(raw.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n"))
                return
        self.text_edit.setPlainText(self.DEFAULT_NOTICE)

    def save_notice(self):
        content = self.text_edit.toPlainText().strip()
        # 줄바꿈을 <br>로 보내서 PWA에서 그대로 표시되도록
        content_html = content.replace("\n", "<br>")
        api = getattr(self.main_win, "api", None) if self.main_win else None
        if api and api.save_notice_api(content_html):
            QMessageBox.information(self, "저장 완료", "공지사항이 백엔드에 저장되었습니다.\nPWA를 새로고침하면 반영됩니다.")
        else:
            QMessageBox.warning(self, "오류", "저장 실패: 백엔드 연결을 확인해 주세요.")


class AdminScreen(QWidget):
    """관리자 등록·수정·삭제·기기 초기화. 테이블: 사번, 이름, 인증."""
    def __init__(self, api, main_win):
        super().__init__()
        self.api = api
        self.main_win = main_win
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        header_layout = QHBoxLayout()
        header = QLabel("관리자")
        header.setObjectName("HeaderTitle")
        header_layout.addWidget(header)
        header_layout.addStretch()
        layout.addLayout(header_layout)

        # 입력 행: 사번, 이름, 등록, 수정, 삭제, 기기 초기화
        input_row = QHBoxLayout()
        input_row.setSpacing(12)
        self.emp_no_input = QLineEdit()
        self.emp_no_input.setPlaceholderText("사번")
        self.emp_no_input.setFixedWidth(180)
        self.emp_no_input.setMinimumHeight(44)
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("이름")
        self.name_input.setFixedWidth(180)
        self.name_input.setMinimumHeight(44)
        input_row.addWidget(self.emp_no_input)
        input_row.addWidget(self.name_input)
        input_row.addSpacing(8)

        self.add_btn = QPushButton("등록")
        self.add_btn.setObjectName("PrimaryBtn")
        self.add_btn.setFixedWidth(100)
        self.add_btn.setMinimumHeight(44)
        self.add_btn.clicked.connect(self.on_add)
        self.edit_btn = QPushButton("수정")
        self.edit_btn.setObjectName("SecondaryBtn")
        self.edit_btn.setFixedWidth(100)
        self.edit_btn.setMinimumHeight(44)
        self.edit_btn.clicked.connect(self.on_edit)
        self.del_btn = QPushButton("삭제")
        self.del_btn.setObjectName("DangerBtn")
        self.del_btn.setFixedWidth(100)
        self.del_btn.setMinimumHeight(44)
        self.del_btn.clicked.connect(self.on_delete)
        self.reset_btn = QPushButton("기기 초기화")
        self.reset_btn.setObjectName("SecondaryBtn")
        self.reset_btn.setFixedWidth(120)
        self.reset_btn.setMinimumHeight(44)
        self.reset_btn.clicked.connect(self.on_reset_device)

        input_row.addWidget(self.add_btn)
        input_row.addWidget(self.edit_btn)
        input_row.addWidget(self.del_btn)
        input_row.addWidget(self.reset_btn)
        input_row.addStretch()
        layout.addLayout(input_row)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["사번", "이름", "인증"])
        setup_standard_table(self.table)
        self.table.itemSelectionChanged.connect(self._on_selection_changed)
        layout.addWidget(self.table, 1)

    def load_data(self):
        self.loader = DataLoader(self.api.get_admins)
        self.loader.finished.connect(self.on_loaded)
        self.loader.start()

    def on_loaded(self, data):
        if not isinstance(data, list):
            return
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(data))
        for i, row in enumerate(data):
            rid = row.get("id")
            self.table.setItem(i, 0, QTableWidgetItem(str(row.get("emp_no") or "")))
            self.table.setItem(i, 1, QTableWidgetItem(str(row.get("name") or "")))
            is_verified = row.get("is_verified", False)
            auth_text = "O" if is_verified else "X"
            item = QTableWidgetItem(auth_text)
            item.setData(Qt.UserRole, rid)
            if is_verified:
                item.setForeground(QColor("#10b981"))
            else:
                item.setForeground(QColor("#ef4444"))
            self.table.setItem(i, 2, item)
        self.table.setSortingEnabled(True)

    def _selected_row_id(self):
        sel = self.table.selectedItems()
        if not sel:
            return None
        row = sel[0].row()
        item = self.table.item(row, 2)
        if item:
            return item.data(Qt.UserRole)
        return None

    def _on_selection_changed(self):
        """테이블 선택 시 사번·이름 입력창에 해당 행 값 표시."""
        rid = self._selected_row_id()
        if rid is None:
            return
        sel = self.table.selectedItems()
        if not sel:
            return
        r = sel[0].row()
        emp_no = (self.table.item(r, 0).text() or "").strip()
        name = (self.table.item(r, 1).text() or "").strip()
        self.emp_no_input.setText(emp_no)
        self.name_input.setText(name)

    def on_add(self):
        emp_no = self.emp_no_input.text().strip()
        name = self.name_input.text().strip()
        if not emp_no:
            QMessageBox.warning(self, "입력 오류", "사번을 입력하세요.")
            return
        if not name:
            QMessageBox.warning(self, "입력 오류", "이름을 입력하세요.")
            return
        success, result = self.api.create_admin(emp_no, name)
        if success:
            QMessageBox.information(self, "등록 완료", "관리자가 등록되었습니다.\n폰에서 최초 로그인 시 비밀번호를 설정해 주세요.")
            self.emp_no_input.clear()
            self.name_input.clear()
            self.load_data()
        else:
            QMessageBox.warning(self, "오류", str(result))

    def on_edit(self):
        rid = self._selected_row_id()
        if rid is None:
            QMessageBox.warning(self, "선택 필요", "수정할 행을 선택하세요.")
            return
        name = self.name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "입력 오류", "이름을 입력하세요.")
            return
        success, detail = self.api.update_admin(rid, name)
        if success:
            QMessageBox.information(self, "수정 완료", "수정되었습니다.")
            self.load_data()
        else:
            QMessageBox.warning(self, "오류", detail or "수정 실패")

    def on_delete(self):
        rid = self._selected_row_id()
        if rid is None:
            QMessageBox.warning(self, "선택 필요", "삭제할 행을 선택하세요.")
            return
        if QMessageBox.question(self, "삭제 확인", "이 관리자를 삭제하시겠습니까?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No) != QMessageBox.StandardButton.Yes:
            return
        if self.api.delete_admin(rid):
            QMessageBox.information(self, "삭제 완료", "삭제되었습니다.")
            self.load_data()
        else:
            QMessageBox.warning(self, "오류", "삭제에 실패했습니다.")

    def on_reset_device(self):
        rid = self._selected_row_id()
        if rid is None:
            QMessageBox.warning(self, "선택 필요", "기기 초기화할 행을 선택하세요.")
            return
        if QMessageBox.question(self, "기기 초기화", "해당 관리자의 기기 인증을 초기화합니다.\n다음 로그인 시 비밀번호를 다시 입력해야 합니다. 계속하시겠습니까?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No) != QMessageBox.StandardButton.Yes:
            return
        if self.api.reset_admin_device(rid):
            QMessageBox.information(self, "초기화 완료", "기기 인증이 초기화되었습니다.")
            self.load_data()
        else:
            QMessageBox.warning(self, "오류", "초기화에 실패했습니다.")


QR_TERMINAL_DIALOG_QSS = """
QDialog#QrTerminalEditDialog {
    background-color: #0f172a;
}
QDialog#QrTerminalEditDialog QLabel {
    color: #f8fafc;
    font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;
    font-weight: bold;
    font-size: 17px;
}
QDialog#QrTerminalEditDialog QLineEdit, QDialog#QrTerminalEditDialog QSpinBox {
    background-color: #1e293b;
    border: 1px solid #475569;
    border-radius: 8px;
    color: #f8fafc;
    padding: 8px 12px;
    font-size: 16px;
    font-weight: bold;
    min-height: 36px;
}
QDialog#QrTerminalEditDialog QLineEdit:disabled, QDialog#QrTerminalEditDialog QSpinBox:disabled {
    background-color: #111b2d;
    color: #64748b;
}
QDialog#QrTerminalEditDialog QSpinBox::up-button, QDialog#QrTerminalEditDialog QSpinBox::down-button {
    background-color: #334155;
    width: 20px;
    border: none;
}
QDialog#QrTerminalEditDialog QCheckBox {
    color: #f8fafc;
    font-weight: bold;
    font-size: 16px;
    font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;
}
QDialog#QrTerminalEditDialog QDialogButtonBox QPushButton {
    background-color: #3b82f6;
    color: white;
    border-radius: 8px;
    padding: 10px 20px;
    font-weight: bold;
    font-size: 16px;
    min-height: 36px;
    min-width: 80px;
    font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;
}
QDialog#QrTerminalEditDialog QDialogButtonBox QPushButton:hover {
    background-color: #2563eb;
}
"""


class QrTerminalEditDialog(QDialog):
    """QR 터미널 추가/수정: 구역명, 스캔 문자열, 프린터·경광등."""
    def __init__(self, parent, api, terminal_id=None):
        super().__init__(parent)
        self.setObjectName("QrTerminalEditDialog")
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.setStyleSheet(QR_TERMINAL_DIALOG_QSS)
        self.api = api
        self.terminal_id = terminal_id
        self.setWindowTitle("QR 터미널 수정" if terminal_id else "QR 터미널 등록")
        self.setMinimumWidth(480)
        form = QFormLayout(self)
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("예: 1번 구역")
        form.addRow("표시 이름:", self.name_edit)
        self.qr_edit = QLineEdit()
        self.qr_edit.setPlaceholderText("스캔 시 들어오는 문자열과 동일 (예: 전체 URL)")
        form.addRow("QR 코드 문자열:", self.qr_edit)

        self.printer_enabled = QCheckBox("식권 프린터 사용")
        form.addRow(self.printer_enabled)
        self.printer_host = QLineEdit()
        self.printer_host.setPlaceholderText("프린터 IP")
        form.addRow("프린터 IP:", self.printer_host)
        self.printer_port = QSpinBox()
        self.printer_port.setRange(1, 65535)
        self.printer_port.setValue(9100)
        form.addRow("프린터 포트:", self.printer_port)

        self.qlight_enabled = QCheckBox("경광등 사용")
        form.addRow(self.qlight_enabled)
        self.qlight_host = QLineEdit()
        self.qlight_host.setPlaceholderText("경광등 IP")
        form.addRow("경광등 IP:", self.qlight_host)
        self.qlight_port = QSpinBox()
        self.qlight_port.setRange(1, 65535)
        self.qlight_port.setValue(20000)
        form.addRow("경광등 포트:", self.qlight_port)

        self.is_active = QCheckBox("사용")
        self.is_active.setChecked(True)
        form.addRow(self.is_active)

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self._on_ok)
        bb.rejected.connect(self.reject)
        form.addRow(bb)

        self.printer_enabled.stateChanged.connect(self._on_printer_toggled)
        self.qlight_enabled.stateChanged.connect(self._on_qlight_toggled)
        self._on_printer_toggled()
        self._on_qlight_toggled()

        if terminal_id and api:
            d = api.get_qr_terminal(terminal_id)
            if d:
                self.name_edit.setText((d.get("name") or "").strip())
                self.qr_edit.setText((d.get("qr_code") or "").strip())
                self.printer_enabled.setChecked(bool(d.get("printer_enabled")))
                self.printer_host.setText((d.get("printer_host") or "").strip())
                self.printer_port.setValue(int(d.get("printer_port") or 9100))
                self.qlight_enabled.setChecked(bool(d.get("qlight_enabled")))
                self.qlight_host.setText((d.get("qlight_host") or "").strip())
                self.qlight_port.setValue(int(d.get("qlight_port") or 20000))
                self.is_active.setChecked(bool(d.get("is_active", True)))
                self._on_printer_toggled()
                self._on_qlight_toggled()

    def _on_printer_toggled(self):
        en = self.printer_enabled.isChecked()
        self.printer_host.setEnabled(en)
        self.printer_port.setEnabled(en)

    def _on_qlight_toggled(self):
        en = self.qlight_enabled.isChecked()
        self.qlight_host.setEnabled(en)
        self.qlight_port.setEnabled(en)

    def _on_ok(self):
        qr = (self.qr_edit.text() or "").strip()
        if not qr:
            QMessageBox.warning(self, "입력 오류", "QR 코드 문자열은 필수입니다.")
            return
        payload = {
            "name": (self.name_edit.text() or "").strip(),
            "qr_code": qr,
            "printer_enabled": self.printer_enabled.isChecked(),
            "printer_host": (self.printer_host.text() or "").strip(),
            "printer_port": self.printer_port.value(),
            "printer_stored_image_number": 1,
            "qlight_enabled": self.qlight_enabled.isChecked(),
            "qlight_host": (self.qlight_host.text() or "").strip(),
            "qlight_port": self.qlight_port.value(),
            "is_active": self.is_active.isChecked(),
        }
        if self.terminal_id:
            ok, err = self.api.update_qr_terminal(self.terminal_id, payload)
        else:
            payload["sort_order"] = 0
            ok, err = self.api.create_qr_terminal(payload)
        if ok:
            self.accept()
        else:
            QMessageBox.warning(self, "오류", str(err or "저장 실패"))


class SettingsScreen(QWidget):
    """설정: QR 터미널(구역별 프린터·경광등)."""
    def __init__(self, api, main_win):
        super().__init__()
        self.api = api
        self.main_win = main_win
        # 바탕은 전역 QSS `QWidget#SettingsScreenRoot`(공지사항 탭 NoticeScreen과 동일 #0f172a). 스크롤만 투명.
        self.setObjectName("SettingsScreenRoot")
        self.setAutoFillBackground(False)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(30, 30, 30, 30)
        outer.setSpacing(12)

        scroll = QScrollArea()
        self._settings_scroll = scroll
        scroll.setObjectName("SettingsScroll")
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setAttribute(Qt.WA_StyledBackground, True)
        scroll.setAutoFillBackground(False)
        _vp = scroll.viewport()
        _vp.setObjectName("SettingsScrollViewport")
        _vp.setAttribute(Qt.WA_StyledBackground, True)
        _vp.setAutoFillBackground(False)

        inner = QWidget()
        self._settings_inner = inner
        inner.setObjectName("SettingsScrollInner")
        inner.setAttribute(Qt.WA_StyledBackground, True)
        inner.setAutoFillBackground(False)
        layout = QVBoxLayout(inner)
        layout.setSpacing(16)

        header = QLabel("설정")
        header.setObjectName("HeaderTitle")
        layout.addWidget(header)

        hint = QLabel(
            "PWA에서 스캔한 문자열이 아래 「QR 문자열」과 정확히 일치할 때만 인증되며, 해당 줄의 프린터·경광등으로 출력됩니다.\n"
            "터미널이 하나도 없으면 서버는 예전 방식(시스템 설정의 허용 QR·기본 장치)을 씁니다. (PC에서는 여기서만 터미널을 관리합니다.)"
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #94a3b8; font-size: 14px; background-color: transparent;")
        layout.addWidget(hint)

        inner.setMinimumWidth(1000)
        _st_adj = _settings_cm_to_px(5.0)
        self._st_table_min_w = max(220, 420 - _st_adj)
        self._st_panel_w = 300 + _st_adj

        layout.addWidget(self._build_printer_section())
        layout.addSpacing(24)
        layout.addWidget(self._build_qlight_section())
        layout.addStretch()

        scroll.setWidget(inner)
        outer.addWidget(scroll)

    def _settings_form_label(self, text):
        lb = QLabel(text)
        lb.setObjectName("InputLabel")
        return lb

    def _settings_parse_port(self, le: QLineEdit) -> int:
        """포트 QLineEdit → 1~65535 정수. 실패 시 경고 후 -1."""
        t = (le.text() or "").strip()
        if not t:
            QMessageBox.warning(self, "입력", "포트를 입력하세요.")
            return -1
        try:
            v = int(t)
            if 1 <= v <= 65535:
                return v
        except ValueError:
            pass
        QMessageBox.warning(self, "입력", "포트는 1~65535 숫자로 입력하세요.")
        return -1

    def _build_printer_section(self):
        """식권 프린터: 테이블(ID, IP, 포트, 연결QR) + 우측 입력 + 추가/수정/삭제."""
        wrap = QWidget()
        v = QVBoxLayout(wrap)
        v.setSpacing(8)
        title = QLabel("식권 프린터 등록")
        title.setStyleSheet("font-weight: bold; font-size: 16px; color: #e2e8f0; background-color: transparent;")
        v.addWidget(title)

        row = QHBoxLayout()
        self.printer_table = QTableWidget(0, 4)
        self.printer_table.setHorizontalHeaderLabels(["ID", "IP", "포트", "연결QR"])
        setup_standard_table(self.printer_table)
        self.printer_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.printer_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        self.printer_table.setColumnWidth(1, 140)
        self.printer_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
        self.printer_table.setColumnWidth(2, 80)
        self.printer_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.printer_table.setMinimumHeight(180)
        self.printer_table.setMinimumWidth(self._st_table_min_w)
        self.printer_table.itemSelectionChanged.connect(self._on_printer_row_selected)
        row.addWidget(self.printer_table, 3)

        panel = QFrame()
        panel.setObjectName("StatCard")
        panel.setFixedWidth(self._st_panel_w)
        pf = QVBoxLayout(panel)
        pf.setSpacing(10)
        pf.addWidget(self._settings_form_label("IP"))
        self.p_ip = QLineEdit()
        self.p_ip.setPlaceholderText("프린터 IP")
        pf.addWidget(self.p_ip)
        pf.addWidget(self._settings_form_label("포트"))
        self.p_port = QLineEdit()
        self.p_port.setPlaceholderText("9100")
        self.p_port.setValidator(QIntValidator(1, 65535, self.p_port))
        self.p_port.setText("9100")
        pf.addWidget(self.p_port)
        pf.addWidget(self._settings_form_label("연결QR"))
        self.p_qr = QLineEdit()
        self.p_qr.setPlaceholderText("PWA 스캔 문자열과 동일")
        pf.addWidget(self.p_qr)
        self.p_active = QCheckBox("사용 (스캔 인증)")
        self.p_active.setChecked(True)
        self.p_active.setStyleSheet("color: #e2e8f0; font-weight: bold;")
        pf.addWidget(self.p_active)
        pbtn = QHBoxLayout()
        pbtn.setSpacing(8)
        self.btn_p_add = QPushButton("추가")
        self.btn_p_add.setObjectName("SettingsActPrimary")
        self.btn_p_add.setFixedHeight(40)
        self.btn_p_add.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.btn_p_add.clicked.connect(self._printer_form_add)
        self.btn_p_edit = QPushButton("수정")
        self.btn_p_edit.setObjectName("SettingsActSecondary")
        self.btn_p_edit.setFixedHeight(40)
        self.btn_p_edit.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.btn_p_edit.clicked.connect(self._printer_form_save)
        self.btn_p_del = QPushButton("삭제")
        self.btn_p_del.setObjectName("SettingsActDanger")
        self.btn_p_del.setFixedHeight(40)
        self.btn_p_del.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.btn_p_del.clicked.connect(self._printer_form_delete)
        pbtn.addWidget(self.btn_p_add, 0)
        pbtn.addWidget(self.btn_p_edit, 0)
        pbtn.addWidget(self.btn_p_del, 0)
        pbtn.addStretch(1)
        pf.addLayout(pbtn)
        pf.addStretch()
        row.addWidget(panel, 1)
        v.addLayout(row)
        self._printer_editing_id = None
        return wrap

    def _build_qlight_section(self):
        """경광등: 테이블 + 우측 입력 + 추가/수정/삭제."""
        wrap = QWidget()
        v = QVBoxLayout(wrap)
        v.setSpacing(8)
        title = QLabel("경광등 등록")
        title.setStyleSheet("font-weight: bold; font-size: 16px; color: #e2e8f0; background-color: transparent;")
        v.addWidget(title)

        row = QHBoxLayout()
        self.qlight_table = QTableWidget(0, 4)
        self.qlight_table.setHorizontalHeaderLabels(["ID", "IP", "포트", "연결QR"])
        setup_standard_table(self.qlight_table)
        self.qlight_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.qlight_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        self.qlight_table.setColumnWidth(1, 140)
        self.qlight_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
        self.qlight_table.setColumnWidth(2, 80)
        self.qlight_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.qlight_table.setMinimumHeight(180)
        self.qlight_table.setMinimumWidth(self._st_table_min_w)
        self.qlight_table.itemSelectionChanged.connect(self._on_qlight_row_selected)
        row.addWidget(self.qlight_table, 3)

        panel = QFrame()
        panel.setObjectName("StatCard")
        panel.setFixedWidth(self._st_panel_w)
        qf = QVBoxLayout(panel)
        qf.setSpacing(10)
        qf.addWidget(self._settings_form_label("IP"))
        self.q_ip = QLineEdit()
        self.q_ip.setPlaceholderText("경광등 IP")
        qf.addWidget(self.q_ip)
        qf.addWidget(self._settings_form_label("포트"))
        self.q_port = QLineEdit()
        self.q_port.setPlaceholderText("20000")
        self.q_port.setValidator(QIntValidator(1, 65535, self.q_port))
        self.q_port.setText("20000")
        qf.addWidget(self.q_port)
        qf.addWidget(self._settings_form_label("연결QR"))
        self.q_qr = QLineEdit()
        self.q_qr.setPlaceholderText("PWA 스캔 문자열과 동일")
        qf.addWidget(self.q_qr)
        self.q_active = QCheckBox("사용 (스캔 인증)")
        self.q_active.setChecked(True)
        self.q_active.setStyleSheet("color: #e2e8f0; font-weight: bold;")
        qf.addWidget(self.q_active)
        qbtn = QHBoxLayout()
        qbtn.setSpacing(8)
        self.btn_q_add = QPushButton("추가")
        self.btn_q_add.setObjectName("SettingsActPrimary")
        self.btn_q_add.setFixedHeight(40)
        self.btn_q_add.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.btn_q_add.clicked.connect(self._qlight_form_add)
        self.btn_q_edit = QPushButton("수정")
        self.btn_q_edit.setObjectName("SettingsActSecondary")
        self.btn_q_edit.setFixedHeight(40)
        self.btn_q_edit.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.btn_q_edit.clicked.connect(self._qlight_form_save)
        self.btn_q_del = QPushButton("삭제")
        self.btn_q_del.setObjectName("SettingsActDanger")
        self.btn_q_del.setFixedHeight(40)
        self.btn_q_del.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.btn_q_del.clicked.connect(self._qlight_form_delete)
        qbtn.addWidget(self.btn_q_add, 0)
        qbtn.addWidget(self.btn_q_edit, 0)
        qbtn.addWidget(self.btn_q_del, 0)
        qbtn.addStretch(1)
        qf.addLayout(qbtn)
        qf.addStretch()
        row.addWidget(panel, 1)
        v.addLayout(row)
        self._qlight_editing_id = None
        return wrap

    def _find_terminal_by_qr(self, qr: str):
        q = (qr or "").strip()
        if not self.api or not q:
            return None
        rows = self.api.list_qr_terminals()
        if not isinstance(rows, list):
            return None
        for row in rows:
            if (row.get("qr_code") or "").strip() == q:
                return row
        return None

    def _on_printer_row_selected(self):
        r = self.printer_table.currentRow()
        if r < 0:
            self._printer_editing_id = None
            return
        it = self.printer_table.item(r, 0)
        if not it:
            return
        tid = it.data(Qt.UserRole)
        self._printer_editing_id = tid
        if not self.api:
            return
        d = self.api.get_qr_terminal(tid)
        if not d:
            return
        self.p_ip.setText((d.get("printer_host") or "").strip())
        self.p_port.setText(str(int(d.get("printer_port") or 9100)))
        self.p_qr.setText((d.get("qr_code") or "").strip())
        self.p_active.setChecked(bool(d.get("is_active", True)))

    def _on_qlight_row_selected(self):
        r = self.qlight_table.currentRow()
        if r < 0:
            self._qlight_editing_id = None
            return
        it = self.qlight_table.item(r, 0)
        if not it:
            return
        tid = it.data(Qt.UserRole)
        self._qlight_editing_id = tid
        if not self.api:
            return
        d = self.api.get_qr_terminal(tid)
        if not d:
            return
        self.q_ip.setText((d.get("qlight_host") or "").strip())
        self.q_port.setText(str(int(d.get("qlight_port") or 20000)))
        self.q_qr.setText((d.get("qr_code") or "").strip())
        self.q_active.setChecked(bool(d.get("is_active", True)))

    def _printer_form_add(self):
        self.printer_table.clearSelection()
        self._printer_editing_id = None
        self.p_ip.clear()
        self.p_port.setText("9100")
        self.p_qr.clear()
        self.p_active.setChecked(True)

    def _qlight_form_add(self):
        self.qlight_table.clearSelection()
        self._qlight_editing_id = None
        self.q_ip.clear()
        self.q_port.setText("20000")
        self.q_qr.clear()
        self.q_active.setChecked(True)

    def _printer_form_save(self):
        if not self.api:
            return
        qr = (self.p_qr.text() or "").strip()
        ip = (self.p_ip.text() or "").strip()
        if not qr:
            QMessageBox.warning(self, "입력", "연결QR은 필수입니다.")
            return
        port = self._settings_parse_port(self.p_port)
        if port < 0:
            return
        if self._printer_editing_id is None:
            existing = self._find_terminal_by_qr(qr)
            if existing:
                ok, err = self.api.update_qr_terminal(
                    existing["id"],
                    {
                        "qr_code": qr,
                        "printer_enabled": True,
                        "printer_host": ip,
                        "printer_port": port,
                        "is_active": self.p_active.isChecked(),
                    },
                )
                msg = "같은 연결QR 터미널에 프린터 정보를 반영했습니다." if ok else str(err or "저장 실패")
            else:
                payload = {
                    "name": "",
                    "qr_code": qr,
                    "printer_enabled": True,
                    "printer_host": ip,
                    "printer_port": port,
                    "printer_stored_image_number": 1,
                    "qlight_enabled": False,
                    "qlight_host": "",
                    "qlight_port": 20000,
                    "is_active": self.p_active.isChecked(),
                    "sort_order": 0,
                }
                ok, err = self.api.create_qr_terminal(payload)
                msg = "등록되었습니다." if ok else str(err or "등록 실패")
            if ok:
                QMessageBox.information(self, "완료", msg)
                self.refresh_terminals()
                self._printer_form_add()
            else:
                QMessageBox.warning(self, "오류", msg)
            return
        ok, err = self.api.update_qr_terminal(
            self._printer_editing_id,
            {
                "qr_code": qr,
                "printer_enabled": True,
                "printer_host": ip,
                "printer_port": port,
                "is_active": self.p_active.isChecked(),
            },
        )
        if ok:
            QMessageBox.information(self, "완료", "수정되었습니다.")
            self.refresh_terminals()
        else:
            QMessageBox.warning(self, "오류", str(err or "수정 실패"))

    def _printer_form_delete(self):
        tid = self._printer_editing_id
        if tid is None:
            QMessageBox.warning(self, "선택", "삭제할 행을 선택하거나 추가 후 목록에서 선택하세요.")
            return
        if QMessageBox.question(
            self,
            "삭제",
            "이 터미널(ID %s)을 삭제하면 같은 연결QR의 경광등 설정도 함께 삭제됩니다. 계속할까요?" % tid,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        if self.api and self.api.delete_qr_terminal(tid):
            QMessageBox.information(self, "완료", "삭제되었습니다.")
            self.refresh_terminals()
            self._printer_form_add()
        else:
            QMessageBox.warning(self, "오류", "삭제에 실패했습니다.")

    def _qlight_form_save(self):
        if not self.api:
            return
        qr = (self.q_qr.text() or "").strip()
        ip = (self.q_ip.text() or "").strip()
        if not qr:
            QMessageBox.warning(self, "입력", "연결QR은 필수입니다.")
            return
        port = self._settings_parse_port(self.q_port)
        if port < 0:
            return
        if self._qlight_editing_id is None:
            existing = self._find_terminal_by_qr(qr)
            if existing:
                ok, err = self.api.update_qr_terminal(
                    existing["id"],
                    {
                        "qr_code": qr,
                        "qlight_enabled": True,
                        "qlight_host": ip,
                        "qlight_port": port,
                        "is_active": self.q_active.isChecked(),
                    },
                )
                msg = "같은 연결QR 터미널에 경광등 정보를 반영했습니다." if ok else str(err or "저장 실패")
            else:
                payload = {
                    "name": "",
                    "qr_code": qr,
                    "printer_enabled": False,
                    "printer_host": "",
                    "printer_port": 9100,
                    "printer_stored_image_number": 1,
                    "qlight_enabled": True,
                    "qlight_host": ip,
                    "qlight_port": port,
                    "is_active": self.q_active.isChecked(),
                    "sort_order": 0,
                }
                ok, err = self.api.create_qr_terminal(payload)
                msg = "등록되었습니다." if ok else str(err or "등록 실패")
            if ok:
                QMessageBox.information(self, "완료", msg)
                self.refresh_terminals()
                self._qlight_form_add()
            else:
                QMessageBox.warning(self, "오류", msg)
            return
        ok, err = self.api.update_qr_terminal(
            self._qlight_editing_id,
            {
                "qr_code": qr,
                "qlight_enabled": True,
                "qlight_host": ip,
                "qlight_port": port,
                "is_active": self.q_active.isChecked(),
            },
        )
        if ok:
            QMessageBox.information(self, "완료", "수정되었습니다.")
            self.refresh_terminals()
        else:
            QMessageBox.warning(self, "오류", str(err or "수정 실패"))

    def _qlight_form_delete(self):
        tid = self._qlight_editing_id
        if tid is None:
            QMessageBox.warning(self, "선택", "삭제할 행을 선택하세요.")
            return
        if QMessageBox.question(
            self,
            "삭제",
            "이 터미널(ID %s)을 삭제하면 같은 연결QR의 프린터 설정도 함께 삭제됩니다. 계속할까요?" % tid,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        if self.api and self.api.delete_qr_terminal(tid):
            QMessageBox.information(self, "완료", "삭제되었습니다.")
            self.refresh_terminals()
            self._qlight_form_add()
        else:
            QMessageBox.warning(self, "오류", "삭제에 실패했습니다.")

    def refresh_terminals(self):
        self.printer_table.setRowCount(0)
        self.qlight_table.setRowCount(0)
        if not self.api:
            return
        rows = self.api.list_qr_terminals()
        if not isinstance(rows, list):
            return
        self.printer_table.setRowCount(len(rows))
        self.qlight_table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            tid = r.get("id")
            id_item = QTableWidgetItem(str(tid))
            id_item.setData(Qt.UserRole, tid)
            qr = (r.get("qr_code") or "").strip()
            ph = (r.get("printer_host") or "").strip()
            pp = r.get("printer_port")
            self.printer_table.setItem(i, 0, id_item)
            self.printer_table.setItem(i, 1, QTableWidgetItem(ph if r.get("printer_enabled") else ""))
            self.printer_table.setItem(i, 2, QTableWidgetItem(str(pp or "") if r.get("printer_enabled") else ""))
            self.printer_table.setItem(i, 3, QTableWidgetItem(qr))

            id_item2 = QTableWidgetItem(str(tid))
            id_item2.setData(Qt.UserRole, tid)
            qh = (r.get("qlight_host") or "").strip()
            qp = r.get("qlight_port")
            self.qlight_table.setItem(i, 0, id_item2)
            self.qlight_table.setItem(i, 1, QTableWidgetItem(qh if r.get("qlight_enabled") else ""))
            self.qlight_table.setItem(i, 2, QTableWidgetItem(str(qp or "") if r.get("qlight_enabled") else ""))
            self.qlight_table.setItem(i, 3, QTableWidgetItem(qr))

    def load_data(self):
        self.refresh_terminals()


class MainWindow(QMainWindow):
    def __init__(self, api=None):
        super().__init__()
        self.api = api if api is not None else APIClient()
        self.companies_data = []
        self.departments_data = []
        self.setWindowTitle("Meal Auth - Admin Management System")
        # 화면보다 큰 최소 크기 요구로 setGeometry 경고 나오지 않도록 상한 설정
        screen = QApplication.primaryScreen()
        if screen:
            available = screen.availableGeometry()
            max_w, max_h = available.width(), available.height()
            self.setMinimumSize(min(1024, max_w), min(700, max_h))
        self.resize(1280, 850)
        # 화면 가용 영역 안으로 초기 크기만 제한 (이후 사용자가 창 크기 조절 가능)
        if screen:
            available = screen.availableGeometry()
            w = min(self.width(), available.width())
            h = min(self.height(), available.height())
            self.resize(w, h)
        self.setStyleSheet(QSS)
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        self.sidebar = QWidget()
        self.sidebar.setObjectName("Sidebar")
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        logo = QLabel("Meal Admin")
        logo.setObjectName("Logo")
        sidebar_layout.addWidget(logo)
        self.nav_btns = []
        menus = [
            ("대시보드", 0), ("회사 관리", 1), ("부서 관리", 2),
            ("사원 관리", 3), ("원시 데이터", 4), ("식사 정책", 5), ("보고서", 6), ("공지사항", 7), ("관리자", 8), ("설정", 9)
        ]
        for name, idx in menus:
            btn = QPushButton(name)
            btn.setObjectName("MenuBtn")
            btn.setProperty("active", "false")
            btn.clicked.connect(lambda chk, i=idx: self.switch_screen(i))
            sidebar_layout.addWidget(btn)
            self.nav_btns.append(btn)
        sidebar_layout.addStretch()
        main_layout.addWidget(self.sidebar)
        content_container = QWidget()
        content_container.setObjectName("ContentArea")
        self.content_layout = QVBoxLayout(content_container)
        self.stack = QStackedWidget()
        self.dashboard = DashboardScreen()
        self.companies = CompanyScreen(self.api, self)
        self.departments = DepartmentScreen(self.api, self)
        self.employees = EmployeeScreen(self.api, self)
        self.raw_data = RawDataScreen(self.api, self)
        self.policies = PolicyScreen(self.api, self)
        self.reports = ReportScreen(self.api, self)
        self.notice_screen = NoticeScreen(self)
        self.admin_screen = AdminScreen(self.api, self)
        self.settings_screen = SettingsScreen(self.api, self)
        self.stack.addWidget(self.dashboard)
        self.stack.addWidget(self.companies)
        self.stack.addWidget(self.departments)
        self.stack.addWidget(self.employees)
        self.stack.addWidget(self.raw_data)
        self.stack.addWidget(self.policies)
        self.stack.addWidget(self.reports)
        self.stack.addWidget(self.notice_screen)
        self.stack.addWidget(self.admin_screen)
        self.stack.addWidget(self.settings_screen)
        self.content_layout.addWidget(self.stack)
        main_layout.addWidget(content_container)
        # QWidget은 기본적으로 QSS 배경을 안 칠함 → Windows에서 탭 바탕이 흰색으로 보임. WA_StyledBackground 필요.
        for _i in range(self.stack.count()):
            _w = self.stack.widget(_i)
            if _w is not None:
                _w.setAttribute(Qt.WA_StyledBackground, True)
                _w.setAutoFillBackground(False)
        self.switch_screen(0)
        self.on_company_changed() # Initial load of companies
        self.refresh_stats()
        
        # Obsolete Auto-refresh timer removed in favor of WebSockets
        
        # Start WebSocket Client
        self.ws_client = WSClient()
        self.ws_client.message_received.connect(self.on_ws_message)
        self.ws_client.start()

    def resizeEvent(self, event):
        """창 크기가 화면 가용 영역을 넘지 않도록 제한 (setGeometry 경고 방지)"""
        screen = QApplication.primaryScreen()
        if screen:
            available = screen.availableGeometry()
            w = event.size().width()
            h = event.size().height()
            if w > available.width() or h > available.height():
                w = min(w, available.width())
                h = min(h, available.height())
                event = QResizeEvent(QSize(w, h), event.oldSize())
        super().resizeEvent(event)

    def on_ws_message(self, data):
        msg_type = data.get("type")
        if msg_type:
            print(f"[WS] message type={msg_type}")
        if msg_type in ["USER_VERIFIED", "MEAL_LOG_CREATED", "STATS_REFRESH"]:
            self.refresh_stats()  # refresh active screen and dashboard
        # PC 앱에서 프린터·경광등 신호 전송 (QR 인증 등 MEAL_LOG_CREATED만, 수동 등록은 STATS_REFRESH)
        if msg_type == "MEAL_LOG_CREATED":
            self._trigger_devices_from_meal_data(data.get("data") or {})

    def _trigger_devices_from_meal_data(self, meal_data: dict):
        """MEAL_LOG_CREATED 수신 시 설정에 따라 PC에서 프린터·경광등 신호 전송."""
        if not meal_data:
            print("[DEVICE] skip: empty meal_data")
            return
        if (meal_data.get("path") or "").upper() == "MANUAL":
            return
        device = meal_data.get("device")
        if not device and self.api:
            device = self.api.get_device_settings()
        if not device:
            print("[DEVICE] skip: no device settings")
            return
        print(
            "[DEVICE] settings "
            f"terminal_id={device.get('terminal_id')} "
            f"printer_enabled={bool(device.get('printer_enabled'))} "
            f"qlight_enabled={bool(device.get('qlight_enabled'))} "
            f"qlight_host={(device.get('qlight_host') or '').strip()} "
            f"qlight_port={int(device.get('qlight_port') or 20000)}"
        )
        if not (device.get("printer_enabled") or device.get("qlight_enabled")):
            print("[DEVICE] skip: printer/qlight both disabled")
            return
        t = threading.Thread(target=_run_print_and_qlight, args=(meal_data, device), daemon=True)
        t.start()

    def on_company_changed(self):
        self.company_sync_loader = DataLoader(self.api.get_companies)
        self.company_sync_loader.finished.connect(self.on_company_sync_finished)
        self.company_sync_loader.start()

    def on_company_sync_finished(self, data):
        if isinstance(data, list):
            self.companies_data = data
            self.departments.update_company_combo(self.companies_data)
            self.employees.update_company_combo(self.companies_data)
            self.sync_departments()

    def sync_departments(self):
        self.dept_sync_loader = DataLoader(self.api.get_departments)
        self.dept_sync_loader.finished.connect(self.on_dept_sync_finished)
        self.dept_sync_loader.start()

    def on_dept_sync_finished(self, data):
        if isinstance(data, list):
            self.departments_data = data
            if self.stack.currentIndex() == 2:
                self.departments.load_data()
            self.employees.update_dept_combo()
            self.reports._refresh_dept_sub_combo()

    def refresh_stats(self):
        self.refresh_active_screen()
        if hasattr(self, 'stat_loader') and self.stat_loader.isRunning():
            return
        self.stat_loader = DataLoader(self.api.get_stats)
        self.stat_loader.finished.connect(self.display_stats)
        self.stat_loader.start()
        
    def refresh_active_screen(self):
        idx = self.stack.currentIndex()
        if idx == 3: # EmployeeScreen
            if not hasattr(self.employees, 'loader') or not self.employees.loader.isRunning():
                self.employees.load_data()
        elif idx == 4: # RawDataScreen
            pass # Explicit search only
    def display_stats(self, stats):
        if stats and isinstance(stats, dict):
            # Dynamic stats update
            self.dashboard.update_stats(stats)
            # 테이블도 서버가 통계 낸 날짜로 조회 (타임존/날짜 불일치 방지)
            self.load_recent_logs(stats.get("date"))

    def load_recent_logs(self, date_value=None):
        if hasattr(self, 'recent_loader') and self.recent_loader.isRunning():
            return
        # 서버 stats의 date 사용 (없으면 한국 로컬(KST) 오늘)
        if date_value:
            if hasattr(date_value, 'strftime'):
                day_str = date_value.strftime("%Y-%m-%d")
            else:
                day_str = str(date_value)[:10]
        else:
            kst = timezone(timedelta(hours=9))
            day_str = datetime.now(kst).date().strftime("%Y-%m-%d")
        self.recent_loader = DataLoader(self.api.get_raw_data, "", day_str, day_str)
        self.recent_loader.finished.connect(self.on_recent_logs_finished)
        self.recent_loader.start()

    def on_recent_logs_finished(self, data):
        if isinstance(data, tuple) and len(data) >= 2 and data[0]:
            data = data[1]
        if isinstance(data, list):
            # Show all logs for today on dashboard
            self.dashboard.update_recent(data)
    def switch_screen(self, idx):
        self.stack.setCurrentIndex(idx)
        for i, btn in enumerate(self.nav_btns):
            btn.setProperty("active", "true" if i == idx else "false")
            btn.style().unpolish(btn)
            btn.style().polish(btn)
        if idx == 1: self.companies.load_data()
        if idx == 2: self.departments.load_data()  # refresh table when opening tab (filters by selected company)
        if idx == 3:
            self.employees.update_dept_combo()  # 부서 콤보 갱신 (회사 1개 기준)
            self.employees.load_data()
        if idx == 4: pass # Explicit search only
        if idx == 5: self.policies.load_data()
        if idx == 6:
            self.reports._refresh_dept_sub_combo()
            self.reports.load_data()
        if idx == 7:
            self.notice_screen.load_notice()
        if idx == 8:
            self.admin_screen.load_data()
        if idx == 9:
            self.settings_screen.load_data()

    def closeEvent(self, event):
        if hasattr(self, 'ws_client'):
            self.ws_client.stop()
        self.api.close()
        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    if PC_APP_SKIP_LOGIN:
        api = APIClient(token=None)
    else:
        login_dialog = AdminLoginDialog()
        if login_dialog.exec_() != QDialog.Accepted or not getattr(login_dialog, "token", None):
            sys.exit(0)
        api = APIClient(token=login_dialog.token)
    window = MainWindow(api=api)
    window.show()
    sys.exit(app.exec_())
