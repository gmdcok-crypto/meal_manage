# -*- coding: utf-8 -*-
"""엑셀 일괄 등록용 양식 시트 생성. 실행: python create_employee_excel_template.py"""
import openpyxl
from openpyxl.styles import Font, Alignment

HEADERS = ["사번", "성명", "부서명"]
EXAMPLE_ROW = ["2024001", "홍길동", "개발팀"]
OUTPUT_FILE = "사원_일괄등록_양식.xlsx"

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "사원목록"

    # 헤더
    for col, name in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # 예시 행 1줄
    for col, val in enumerate(EXAMPLE_ROW, 1):
        ws.cell(row=2, column=col, value=val)

    # 빈 행 (사용자가 입력할 공간)
    for r in range(3, 8):
        for c in range(1, 4):
            ws.cell(row=r, column=c, value="")

    # 컬럼 너비
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14

    wb.save(OUTPUT_FILE)
    print(f"생성됨: {OUTPUT_FILE}")
    print("필수 컬럼: 사번, 성명, 부서명 (1행 헤더 유지)")

if __name__ == "__main__":
    main()
